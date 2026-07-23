# ==========================================
# [PART_1_START] - Framework Imports & Theme Styling
# ==========================================
import streamlit as st
import sqlite3
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
import os
import base64
from PIL import Image

# पीडीएफ जनरेशन के लिए रिपोर्टलैब फ्रेमवर्क
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Helper helper to safely convert local BMP files to baseline inline HTML strings
def get_local_bmp_base64(file_name):
    if os.path.exists(file_name):
        try:
            with open(file_name, "rb") as image_file:
                encoded_string = base64.b64encode(image_file.read()).decode()
            return f"data:image/bmp;base64,{encoded_string}"
        except Exception:
            return None
    return None

# =========================================================================
# ⏰ GLOBAL LAPTOP CLOCK SYNCHRONIZATION ENGINE
# =========================================================================
def get_laptop_time():
    """Yeh function hardcoded backend server times ke badle direct aapke laptop ka live exact time database entries me force karta hai."""
    ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
    return ist_now.strftime("%Y-%m-%d %H:%M:%S")

# =========================================================================
# 1. FIXED PAGE CONFIGURATION FOR FULL-WIDTH VIEW
# =========================================================================
st.set_page_config(page_title="Mannat Wealth Master", page_icon="📊", layout="wide")

# -------------------------------------------------------------------------
# 🔥 STEP 1: PREMIUM SYMMETRICAL DIVINE BANNER (LOCAL BMP DISK DISCOVERY LAYER)
# -------------------------------------------------------------------------
col_balaji, col_main_title, col_shyam = st.columns([0.15, 0.7, 0.15])

# Fetch data representations directly from local disk execution paths
balaji_b64 = get_local_bmp_base64("balaji.bmp")
shyam_b64 = get_local_bmp_base64("shyam.bmp")

with col_balaji:
    if balaji_b64:
        st.markdown(f"""
            <div style='text-align: center;'>
                <img src='{balaji_b64}' width='100' style='border-radius: 8px; filter: drop-shadow(0px 4px 8px rgba(250, 140, 22, 0.3));'>
                <div style='font-size: 11px; font-weight: bold; color: #D46B08; margin-top: 5px;'>BALAJI SARKAR</div>
            </div>
        """, unsafe_allow_html=True)
    else:
        # Graceful decorative icon fallback if file is not found in directory path
        st.markdown("""
            <div style='text-align: center;'>
                <span style='font-size: 55px; line-height: 1;'>🔱</span>
                <div style='font-size: 11px; font-weight: bold; color: #D46B08; margin-top: 2px;'>BALAJI SARKAR</div>
            </div>
        """, unsafe_allow_html=True)

with col_main_title:
    st.markdown("""
        <div style='text-align: center; margin-top: 5px;'>
            <h1 style='color: #1E3A8A; font-family: sans-serif; margin-bottom: 0px; padding-bottom: 0px; font-weight: bold; letter-spacing: 0.5px;'>
                MANNAT WEALTH
            </h1>
            <p style='font-size: 14px; color: #6B7280; margin-top: 5px; padding-top: 0px; font-weight: 500;'>
                Ultimate Multi-Client Algorithmic Trading Ledger & Cockpit System (MW Version 20.4)
            </p>
        </div>
    """, unsafe_allow_html=True)

with col_shyam:
    if shyam_b64:
        st.markdown(f"""
            <div style='text-align: center;'>
                <img src='{shyam_b64}' width='100' style='border-radius: 8px; filter: drop-shadow(0px 4px 8px rgba(214, 40, 40, 0.3));'>
                <div style='font-size: 11px; font-weight: bold; color: #B71C1C; margin-top: 5px;'>SHYAM BABA</div>
            </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
            <div style='text-align: center;'>
                <span style='font-size: 55px; line-height: 1;'>🪔</span>
                <div style='font-size: 11px; font-weight: bold; color: #B71C1C; margin-top: 2px;'>SHYAM BABA</div>
            </div>
        """, unsafe_allow_html=True)

st.markdown("---")

# =========================================================================
# 🛡️ SECURITY IMPLEMENTATION: HARDCODED BYPASS-PROOF EXTENSION MODULE
# =========================================================================
# 🔥 BHAI, JAB BHI DATE EXTEND KARNI HO, IS LINE KI DATE BADAL KAR NAYI FILE RE-COMPILE (.pyc) KARKE BHEJ DENA!
EXPIRY_DATE_STR = "2026-09-10" 
EXPIRY_DATE = datetime.strptime(EXPIRY_DATE_STR, "%Y-%m-%d").date()
CURRENT_DATE = datetime.now().date()

# 🔔 DYNAMIC 10-DAY BEFORE PRE-ALERT DISPATCH TRIGGER
days_remaining = (EXPIRY_DATE - CURRENT_DATE).days

if 0 <= days_remaining <= 10:
    st.markdown(
        f"""
        <div style="background-color:#fff3cd; padding:15px; border-radius:8px; border-left:5px solid #ffc107; margin-bottom:20px;">
            <strong style="color:#856404; font-size:16px;">⚠️ Subscription Expiry Warning!</strong>
            <p style="color:#856404; margin:5px 0 0 0; font-size:14px;">
                Aapka software subscription block agle <b>{days_remaining} din</b> mein khatam hone wala hai ({EXPIRY_DATE_STR}). 
                Kripya access continuous rakhne ke liye system administrator se sampark karein.
            </p>
        </div>
        """, 
        unsafe_allow_html=True
    )

if CURRENT_DATE > EXPIRY_DATE:
    st.error("❌ SUBSCRIPTION EXPIRED / RENEWAL REQUIRED")
    st.markdown(
        f"""
        <div style="background-color:#ffe6e6; padding:20px; border-radius:10px; border:2px solid #ff4d4d;">
            <h2 style="color:#cc0000; margin-top:0;">⚠️ Access Suspended</h2>
            <p style="font-size:16px; color:#333;">
                Mannat Wealth Cockpit application ka subscription block <b>{EXPIRY_DATE_STR}</b> ko end ho chuka hai.
            </p>
            <p style="font-size:15px; color:#555;">
                System operational state aur metrics access ko restore karne ke liye system administrator se sampark karein aur renew karwayen.
            </p>
            <hr style="border: 0; border-top: 1px solid #ff9999;">
            <p style="font-size:12px; color:#888; margin-bottom:0;">
                <b>System UUID:</b> MW-V20-ULTIMATE-LOCK | <b>Status Code:</b> 402_RENEWAL_PENDING
            </p>
        </div>
        """, 
        unsafe_allow_html=True
    )
    st.stop()
# ==========================================
# [PART_1_END]
# ==========================================
# ==========================================
# [PART_2_START] - Login Gate & Pastel Card CSS CSS
# ==========================================
def check_password():
    """यदि यूजर सही क्रेडेंशियल डालता है तो True रिटर्न करता है।"""
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False

    if st.session_state["authenticated"]:
        return True

    st.markdown("<h2 style='text-align: center; color: #0d47a1;'>🔐 JAI SHRI BALAJI </h2>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c2:
        with st.form("login_form_master"):
            username = st.text_input("Username / Login ID", placeholder="Enter ID")
            password = st.text_input("Password", type="password", placeholder="Enter Password")
            submit = st.form_submit_button("🔓 Log In & Open Cockpit", use_container_width=True)
            
            if submit:
                if username == "admin" and password == "salasar@2026":
                    st.session_state["authenticated"] = True
                    st.success("Access Granted! Loading system logs...")
                    st.rerun()
                else:
                    st.error("❌ Invalid Login ID or Password. Please try again.")
    return False

if not check_password():
    st.stop()

# =========================================================================
# 3. PREMIUM FRONT PASTEL THEME CARD STYLING
# =========================================================================
st.markdown("""
    <style>
    .metric-card {
        padding: 15px; border-radius: 10px; margin-bottom: 15px;
        box-shadow: 0 4px 10px rgba(0, 0, 0, 0.03); border: 1px solid #eaeaea;
    }
    .pastel-green { background-color: #e2f4e9; color: #1e5631; border-left: 5px solid #28a745; }
    .pastel-red { background-color: #fbe7e9; color: #842029; border-left: 5px solid #dc3545; }
    .pastel-yellow { background-color: #fff8e1; color: #b27a00; border-left: 5px solid #ffc107; }
    .pastel-blue { background-color: #e3f2fd; color: #0d47a1; border-left: 5px solid #1e88e5; }
    </style>
""", unsafe_allow_html=True)
# ==========================================
# [PART_2_END]
# ==========================================
# ==========================================
# [PART_3_START] - Automated Anti-Crash Backup Engine & Initialisation Layer
# ==========================================
def execute_database_daily_backup():
    """Yeh function application startup par pichli database file ka local automatic backup banata hai."""
    import shutil
    import os
    primary_db_name = 'salasar_wealth_v19_ultimate.db'
    
    if os.path.exists(primary_db_name):
        try:
            current_date_str = datetime.now().strftime("%Y-%m-%d")
            backup_folder = "salasar_db_vault"
            
            if not os.path.exists(backup_folder):
                os.makedirs(backup_folder)
                
            backup_filename = os.path.join(backup_folder, f"salasar_backup_{current_date_str}.db")
            
            if not os.path.exists(backup_filename):
                shutil.copy2(primary_db_name, backup_filename)
                
                all_backups = sorted([os.path.join(backup_folder, f) for f in os.listdir(backup_folder) if f.endswith('.db')])
                if len(all_backups) > 14:
                    for old_backup_file in all_backups[:-14]:
                        os.remove(old_backup_file)
        except Exception as e:
            print(f"Backup operation skipped/failed: {str(e)}")

def init_db():
    execute_database_daily_backup() # 🔥 First execution checkpoint triggers backup save
    conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS trades (
            id INTEGER PRIMARY KEY AUTOINCREMENT, client_name TEXT, week_block TEXT, exchange TEXT, script_name TEXT,
            selected_expiry TEXT, action_type TEXT, buy_qty INTEGER, buy_price REAL, sell_qty INTEGER, sell_price REAL,
            turnover REAL, brokerage REAL, manual_pnl REAL, status TEXT DEFAULT 'CARRY FORWARD', timestamp TEXT
        )''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_settings (
            client_name TEXT PRIMARY KEY, opening_balance REAL, brokerage_nse REAL, brokerage_mcx REAL, brokerage_gifty REAL, 
            expiry_default TEXT, expiry_optional TEXT, brokerage_type TEXT DEFAULT 'Per Crore', per_lot_rate REAL DEFAULT 0.0,
            whatsapp_phone TEXT DEFAULT ''
        )''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS global_market_prices (
            symbol TEXT PRIMARY KEY, closing_price REAL, updated_at TEXT
        )''')
        
    # 🔥 MULTI-MAPPING UPGRADE LAYER: Creating a separate mapping database table safely inside open connection
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sub_broker_mappings (
            client_name TEXT,
            sub_broker_tag TEXT,
            timestamp TEXT,
            PRIMARY KEY (client_name, sub_broker_tag)
        )''')
        
    # Automatic historical data migration matrix rule
    try:
        cursor.execute("PRAGMA table_info(master_clients)")
        columns_mc = [c[1] for c in cursor.fetchall()]
        if "sub_broker_tag" in columns_mc:
            cursor.execute("SELECT client_name, sub_broker_tag, timestamp FROM master_clients WHERE sub_broker_tag IS NOT NULL AND sub_broker_tag != 'None'")
            old_rows = cursor.fetchall()
            for c_n, sb_t, t_s in old_rows:
                cursor.execute("INSERT OR IGNORE INTO sub_broker_mappings (client_name, sub_broker_tag, timestamp) VALUES (?, ?, ?)", (c_n, sb_t, t_s))
    except Exception as err:
        print(f"Migration skipped: {str(err)}")
        
    try:
        cursor.execute("SELECT client_name, expiry_default FROM client_settings")
        settings_rows = cursor.fetchall()
        for c_profile, def_exp in settings_rows:
            if def_exp and str(def_exp).strip():
                cursor.execute("""
                    UPDATE trades 
                    SET selected_expiry = ? 
                    WHERE client_name = ? 
                    AND (selected_expiry IS NULL OR selected_expiry = '' OR selected_expiry = 'None' OR selected_expiry = '()')
                """, (str(def_exp).strip(), c_profile))
    except Exception as e: print(f"Database alignment skipped: {str(e)}")

    try: cursor.execute("ALTER TABLE client_settings ADD COLUMN brokerage_type TEXT DEFAULT 'Per Crore'")
    except: pass
    try: cursor.execute("ALTER TABLE client_settings ADD COLUMN per_lot_rate REAL DEFAULT 0.0")
    except: pass
    try: cursor.execute("ALTER TABLE client_settings ADD COLUMN whatsapp_phone TEXT DEFAULT ''")
    except: pass
# ==========================================
# [PART_3_END]
# ==========================================
# ==========================================
# [PART_4_START] - Operational Memory Loading & Profile Setup
# ==========================================
    # 🔥 RE-ARRANGED ORDER: Table queries execute smoothly under single active initialization stream
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS cash_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT, client_name TEXT, week_block TEXT, tx_type TEXT, amount REAL, remarks TEXT, timestamp TEXT
        )''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS operational_weeks (
            week_name TEXT PRIMARY KEY, timestamp TEXT
        )''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS master_clients (
            client_name TEXT PRIMARY KEY, timestamp TEXT
        )''')
    current_laptop_time = get_laptop_time()
    default_blocks = ["15-19 June, 2026", "22-26 June, 2026", "29 June-3 July, 2026"]
    for block in default_blocks:
        cursor.execute("INSERT OR IGNORE INTO operational_weeks (week_name, timestamp) VALUES (?, ?)", (block, current_laptop_time))
    cursor.execute("SELECT COUNT(*) FROM master_clients")
    if cursor.fetchone()[0] == 0:
        default_clients = ['Pj Nse', 'Pj Mcx', 'Pj Sgx', 'DG001', 'Dg002', 'Dg003', 'RG', 'Master 2', 'Jitneder', 'Tony']
        for client in default_clients:
            cursor.execute("INSERT OR IGNORE INTO master_clients (client_name, timestamp) VALUES (?, ?)", (client, current_laptop_time))
    conn.commit()
    conn.close()

init_db()

def get_global_price(symbol):
    conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    cursor = conn.cursor()
    cursor.execute("SELECT closing_price FROM global_market_prices WHERE symbol = ?", (str(symbol).lower().strip(),))
    row = cursor.fetchone()
    conn.close()
    return float(row[0]) if row and row is not None else 0.0

def update_global_price(symbol, price):
    conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    cursor = conn.cursor()
    laptop_time = get_laptop_time()
    cursor.execute('''
        INSERT INTO global_market_prices (symbol, closing_price, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(symbol) DO UPDATE SET closing_price=excluded.closing_price, updated_at=excluded.updated_at
    ''', (str(symbol).lower().strip(), float(price), laptop_time))
    conn.commit()
    conn.close()

def load_clients():
    conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    cursor = conn.cursor()
    cursor.execute("SELECT client_name FROM master_clients ORDER BY timestamp ASC")
    rows = cursor.fetchall()
    conn.close()
    return [str(r[0]) for r in rows] if rows else ['Pj Nse']

def load_stored_weeks():
    conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    cursor = conn.cursor()
    cursor.execute("SELECT week_name FROM operational_weeks ORDER BY timestamp ASC")
    rows = cursor.fetchall()
    conn.close()
    return [str(r[0]) for r in rows] if rows else ["15-19 June, 2026"]

CLIENTS = load_clients()
EXCHANGES = ['NSE', 'MCX', 'GIFTY']
week_options = load_stored_weeks()

if 'active_client_sync' not in st.session_state:
    st.session_state['active_client_sync'] = CLIENTS[0] if CLIENTS else "Pj Nse"

def sync_sidebar_to_form():
    st.session_state['t_cl_fixed_matrix_key'] = st.session_state['sb_cl_selector_final_matrix']

st.sidebar.title("🛠️ Profile & Control Center")
active_client = st.sidebar.selectbox(
    "Select Profile To Configure", CLIENTS, 
    key="sb_cl_selector_final_matrix", on_change=sync_sidebar_to_form
)
# ==========================================
# [PART_4_END]
# ==========================================
# ==========================================
# [PART_5_START] - Sidebar Configuration Forms & Explicit Multi-Expiry Columns Fixed
# ==========================================
conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
cursor = conn.cursor()
# Auto-Repair Engine: Direct structural upgrades inject the missing explicit column blocks into database
try: cursor.execute("ALTER TABLE client_settings ADD COLUMN expiry_3 TEXT DEFAULT '31-Aug-2026'")
except: pass
try: cursor.execute("ALTER TABLE client_settings ADD COLUMN expiry_4 TEXT DEFAULT '15-Jul-2026'")
except: pass
try: cursor.execute("ALTER TABLE client_settings ADD COLUMN expiry_5 TEXT DEFAULT '28-Aug-2026'")
except: pass

cursor.execute("SELECT opening_balance, brokerage_nse, brokerage_mcx, brokerage_gifty, expiry_default, expiry_optional, brokerage_type, per_lot_rate, whatsapp_phone, expiry_3, expiry_4, expiry_5 FROM client_settings WHERE client_name = ?", (active_client,))
saved_client_row = cursor.fetchone()
conn.close()

if saved_client_row and saved_client_row is not None:
    db_cap, db_nse, db_mcx, db_gifty, db_exp1, db_exp2, db_b_type, db_lot_rate, db_phone, db_exp3, db_exp4, db_exp5 = saved_client_row
else:
    db_cap = 0.0
    db_nse, db_mcx, db_gifty = 1000, 1000, 1000
    db_exp1, db_exp2 = "30-Jun-2026", "30-Jul-2026"
    db_b_type, db_lot_rate = "Per Crore", 0.0
    db_phone = ""
    db_exp3, db_exp4, db_exp5 = "31-Aug-2026", "15-Jul-2026", "28-Aug-2026"

opening_balance = st.sidebar.number_input(f"Opening Balance for {active_client} (₹)", value=float(db_cap), key=f"sb_op_bal_{active_client}")
b_type_selection = st.sidebar.selectbox("Brokerage Type", ["Per Crore", "Per Lot"], index=0 if db_b_type == "Per Crore" else 1, key=f"b_type_{active_client}")

if b_type_selection == "Per Crore":
    brokerage_nse = st.sidebar.number_input(f"NSE Rate (Per Crore) (₹)", value=int(db_nse), key=f"sb_nse_r_{active_client}")
    brokerage_mcx = st.sidebar.number_input(f"MCX Rate (Per Crore) (₹)", value=int(db_mcx), key=f"sb_mcx_r_{active_client}")
    brokerage_gifty = st.sidebar.number_input(f"GIFTY Rate (Per Crore) (₹)", value=int(db_gifty), key=f"sb_gifty_r_{active_client}")
    per_lot_rate = 0.0
else:
    per_lot_rate = st.sidebar.number_input(f"Fixed Rate Per Lot (₹)", value=float(db_lot_rate), key=f"sb_lot_r_{active_client}")
    brokerage_nse, brokerage_mcx, brokerage_gifty = 0, 0, 0

# --- SEPARATE DISTINCT EXPIRY COLUMNS GRID ---
default_expiry = st.sidebar.text_input("Default Expiry", value=str(db_exp1), key=f"def_exp_{active_client}")
optional_expiry = st.sidebar.text_input("Optional Expiry", value=str(db_exp2), key=f"opt_exp_{active_client}")

# FIXED LABELS: Clean generic text headers for maximum dynamic asset flexibility
expiry_3_input = st.sidebar.text_input("📅 Expiry 3", value=str(db_exp3 if db_exp3 else "31-Aug-2026"), key=f"exp3_inp_{active_client}")
expiry_4_input = st.sidebar.text_input("📅 Expiry 4", value=str(db_exp4 if db_exp4 else "03-Jul-2026"), key=f"exp4_inp_{active_client}")
expiry_5_input = st.sidebar.text_input("📅 Expiry 5", value=str(db_exp5 if db_exp5 else "28-Aug-2026"), key=f"exp5_inp_{active_client}")

# 🔥 NEW MASTER TOGGLE: Checkbox to push these expiries to all client profiles globally
apply_to_all_clients_global_chk = st.sidebar.checkbox("🔥 Apply this Expiry Schedule to ALL Clients", value=False, key="global_expiry_sync_toggle_chk")

# Consolidate all explicit options directly inside the select box array matrix loop
expiry_options = [default_expiry.strip(), optional_expiry.strip()]
for exp_item in [expiry_3_input, expiry_4_input, expiry_5_input]:
    if exp_item.strip() and exp_item.strip() not in expiry_options:
        expiry_options.append(exp_item.strip())

whatsapp_phone_input = st.sidebar.text_input("🟢 Profile WhatsApp Number", value=str(db_phone if db_phone else ""), placeholder="e.g. 919876543210", key=f"sb_phone_{active_client}")

if st.sidebar.button("💾 Save Client Configuration", use_container_width=True, key=f"sb_save_btn_{active_client}"):
    conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    cursor = conn.cursor()
    
    # Save parameters normally for the selected active profile context first
    cursor.execute("""
        INSERT INTO client_settings (client_name, opening_balance, brokerage_nse, brokerage_mcx, 
        brokerage_gifty, expiry_default, expiry_optional, brokerage_type, per_lot_rate, whatsapp_phone, expiry_3, expiry_4, expiry_5) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) 
        ON CONFLICT(client_name) DO UPDATE SET 
            opening_balance=excluded.opening_balance, brokerage_nse=excluded.brokerage_nse, 
            brokerage_mcx=excluded.brokerage_mcx, brokerage_gifty=excluded.brokerage_gifty, 
            expiry_default=excluded.expiry_default, expiry_optional=excluded.expiry_optional,
            brokerage_type=excluded.brokerage_type, per_lot_rate=excluded.per_lot_rate,
            whatsapp_phone=excluded.whatsapp_phone, expiry_3=excluded.expiry_3, expiry_4=excluded.expiry_4, expiry_5=excluded.expiry_5
    """, (active_client, opening_balance, brokerage_nse, brokerage_mcx, brokerage_gifty, default_expiry, optional_expiry, b_type_selection, per_lot_rate, whatsapp_phone_input.strip(), expiry_3_input.strip(), expiry_4_input.strip(), expiry_5_input.strip()))
    
    # 🔥 CORE SYNC LOGIC ENGINE: If checkbox is active, fire a transactional loop block across all profiles
    if apply_to_all_clients_global_chk:
        for single_profile in CLIENTS:
            if single_profile != active_client:
                # Strictly injects or updates only the expiry columns into other profiles while keeping their balances/rates untouched
                cursor.execute("""
                    INSERT INTO client_settings (client_name, opening_balance, brokerage_nse, brokerage_mcx, brokerage_gifty, expiry_default, expiry_optional, expiry_3, expiry_4, expiry_5)
                    VALUES (?, 0.0, 1000, 1000, 1000, ?, ?, ?, ?, ?)
                    ON CONFLICT(client_name) DO UPDATE SET
                        expiry_default=excluded.expiry_default, expiry_optional=excluded.expiry_optional,
                        expiry_3=excluded.expiry_3, expiry_4=excluded.expiry_4, expiry_5=excluded.expiry_5
                """, (single_profile, default_expiry, optional_expiry, expiry_3_input.strip(), expiry_4_input.strip(), expiry_5_input.strip()))
                
    conn.commit()
    conn.close()
    if apply_to_all_clients_global_chk:
        st.sidebar.success("✨ Master Sync Success! All Commodity Expiries Saved to ALL Clients Globally!")
    else:
        st.sidebar.success("All Commodity Expiries Saved Safely for this profile!")
    st.rerun()
# ==========================================
# [PART_5_END]
# ==========================================
# ==========================================
# [PART_6_START] - Client Add & Sub-Broker Allocation Hub
# ==========================================
st.sidebar.subheader("👤 Manage Clients")
new_client_inp = st.sidebar.text_input("Add New Client Name", key="nc_inp")
if st.sidebar.button("➕ Add Client", use_container_width=True, key="add_cl_master_btn"):
    if new_client_inp.strip():
        current_laptop_time = get_laptop_time()
        conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO master_clients (client_name, timestamp) VALUES (?, ?)", (new_client_inp.strip(), current_laptop_time))
            conn.commit()
            st.sidebar.success(f"Client '{new_client_inp}' Added!")
            st.rerun()
        except sqlite3.IntegrityError: st.sidebar.warning("Client already exists.")
        finally: conn.close()

# --- 🔥 MULTI-MAPPING COCKPIT LAYER ---
st.sidebar.write("---")
st.sidebar.markdown("#### 🔗 Map Client to Sub-Broker")

try:
    conn_sb = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    cursor_sb = conn_sb.cursor()
    cursor_sb.execute("SELECT client_name, sub_broker_tag FROM sub_broker_mappings")
    all_mapping_rows = cursor_sb.fetchall()
    conn_sb.close()
except Exception:
    all_mapping_rows = []

mapping_client_select = st.sidebar.selectbox("Select Target Client", CLIENTS, key="sb_map_client_select_box")

# 🔥 CASE-INSENSITIVE NORMALIZATION VIEW MATRIX LOOP
linked_brokers = [str(r[1]).upper().strip() for r in all_mapping_rows if str(r[0]).strip() == str(mapping_client_select).strip()]
if linked_brokers:
    st.sidebar.info(f"📋 Currently active inside: {', '.join(set(linked_brokers))}")

sub_broker_tag_input = st.sidebar.text_input("Enter Sub-Broker Group Tag", value="", placeholder="e.g. Sarkar-1", key="sb_tag_text_input_field")

if st.sidebar.button("🔗 Lock Sub-Broker Allocation Matrix", use_container_width=True, key="save_sub_broker_mapping_btn"):
    if sub_broker_tag_input.strip():
        # 🔥 FORCE STRIP UPPER UNIFICATION TO PREVENT DUPLICATES GENERATION CAUSING BY SMALL LETTERS
        final_sb_tag = sub_broker_tag_input.strip().upper()
        current_laptop_time = get_laptop_time()
        conn_sb_save = sqlite3.connect('salasar_wealth_v19_ultimate.db')
        cursor_sb_save = conn_sb_save.cursor()
        
        cursor_sb_save.execute("""
            INSERT INTO sub_broker_mappings (client_name, sub_broker_tag, timestamp) 
            VALUES (?, ?, ?)
            ON CONFLICT(client_name, sub_broker_tag) DO UPDATE SET timestamp=excluded.timestamp
        """, (mapping_client_select, final_sb_tag, current_laptop_time))
        conn_sb_save.commit()
        conn_sb_save.close()
        st.sidebar.success(f"✨ Success! '{mapping_client_select}' safely mapped to Sub-Broker: [{final_sb_tag}].")
        st.rerun()
    else:
        st.sidebar.warning("Please type sub-broker tag name.")

# --- 🔥 MULTI-MAPPING MAINTENANCE CONTROLS (RENAME & DELETE) ---
st.sidebar.write("---")
st.sidebar.markdown("#### ⚙️ Edit / Maintenance Sub-Brokers")

# 🔥 UNIQUE UPPERCASE NORMALIZATION ARRAY FOR DROP-DOWN BOX VIEWS
sb_maintenance_list = sorted(list(set([str(r[1]).upper().strip() for r in all_mapping_rows if r and r[1] and str(r[1]).strip() != "None"])))

if sb_maintenance_list:
    target_sb_maintenance = st.sidebar.selectbox("Select Sub-Broker to Edit", sb_maintenance_list, key="sb_maint_select_box")
    sb_rename_new_name = st.sidebar.text_input("Enter New Name for this Broker", value=target_sb_maintenance, key="sb_maint_rename_inp")
    
    col_sb_1, col_sb_2 = st.sidebar.columns(2)
    with col_sb_1:
        if st.button("✏️ Rename Broker", use_container_width=True, key="sb_force_rename_action_btn"):
            if sb_rename_new_name.strip():
                new_broker_name_upper = sb_rename_new_name.strip().upper()
                conn_sb_rn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                cursor_sb_rn = conn_sb_rn.cursor()
                # 🔥 CASE INSENSITIVE BULK UPDATE ENGINE PATCH
                cursor_sb_rn.execute("UPDATE sub_broker_mappings SET sub_broker_tag = ? WHERE UPPER(sub_broker_tag) = UPPER(?)", (new_broker_name_upper, target_sb_maintenance))
                conn_sb_rn.commit()
                conn_sb_rn.close()
                st.sidebar.success("✨ Broker renamed successfully!")
                st.rerun()
    with col_sb_2:
        chk_sb_del = st.checkbox("Confirm Delete Link", key="sb_maint_del_chk_lock")
        if st.button("🗑️ Unmap Client From Broker", use_container_width=True, type="primary", key="sb_force_delete_action_btn"):
            if chk_sb_del:
                conn_sb_del = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                cursor_sb_del = conn_sb_del.cursor()
                # 🔥 CASE INSENSITIVE STRIKE SHIELD LOGIC: Safely wipe out small/capital entries simultaneously 
                cursor_sb_del.execute("DELETE FROM sub_broker_mappings WHERE client_name = ? AND UPPER(sub_broker_tag) = UPPER(?)", (mapping_client_select, target_sb_maintenance))
                conn_sb_del.commit()
                conn_sb_del.close()
                st.sidebar.error("💥 Broker map link cleared safely!")
                st.rerun()
            else:
                st.sidebar.warning("Tick box!")
else:
    st.sidebar.info("No sub-brokers logged yet.")
# ==========================================
# [PART_6_END]
# ==========================================
# ==========================================
# [PART_7_START] - Client Profile Rename, Delete & Operational Weeks Hub
# ==========================================
# --- LIVE CLIENT PROFILE RENAME MATRIX ENGINE ---
st.sidebar.write("---")
st.sidebar.markdown("#### 📝 Rename Existing Client Profile")
rename_target_select = st.sidebar.selectbox("Select Client to Rename", CLIENTS, key="rn_target_sel")
rename_new_name_inp = st.sidebar.text_input("Enter New Name for Profile", value="", placeholder="e.g. Pj", key="rn_new_name_inp")

if st.sidebar.button("⚙️ Force Execute Profile Rename", use_container_width=True, type="secondary", key="execute_rename_btn"):
    old_name_clean = rename_target_select.strip()
    new_name_clean = rename_new_name_inp.strip()
    
    if not new_name_clean:
        st.sidebar.warning("⚠️ Action blocked! New name cannot be empty.")
    elif new_name_clean in CLIENTS:
        st.sidebar.error("❌ Rename failed! This profile name already exists inside system memory.")
    else:
        with st.sidebar.spinner("Re-mapping complete historical database rows..."):
            conn_rn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
            cursor_rn = conn_rn.cursor()
            cursor_rn.execute("UPDATE master_clients SET client_name = ? WHERE client_name = ?", (new_name_clean, old_name_clean))
            cursor_rn.execute("UPDATE client_settings SET client_name = ? WHERE client_name = ?", (new_name_clean, old_name_clean))
            cursor_rn.execute("UPDATE trades SET client_name = ? WHERE client_name = ?", (new_name_clean, old_name_clean))
            cursor_rn.execute("UPDATE cash_transactions SET client_name = ? WHERE client_name = ?", (new_name_clean, old_name_clean))
            conn_rn.commit()
            conn_rn.close()
            st.sidebar.success(f"✨ Success! '{old_name_clean}' permanently renamed to '{new_name_clean}'.")
            st.rerun()

st.sidebar.write("---")
del_client_select = st.sidebar.selectbox("Select Client to Delete", CLIENTS, key="dc_sel")
confirm_del_cl = st.sidebar.checkbox("Confirm Client Deletion", key="c_del_cl")
if st.sidebar.button("🗑️ Delete Selected Client", use_container_width=True, type="primary", key="del_cl_master_btn"):
    if confirm_del_cl:
        conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
        cursor = conn.cursor()
        cursor.execute("DELETE FROM master_clients WHERE client_name = ?", (del_client_select,))
        cursor.execute("DELETE FROM client_settings WHERE client_name = ?", (del_client_select,))
        cursor.execute("DELETE FROM trades WHERE client_name = ?", (del_client_select,))
        cursor.execute("DELETE FROM cash_transactions WHERE client_name = ?", (del_client_select,))
        conn.commit()
        conn.close()
        st.sidebar.error("Client Wiped Safely!")
        st.rerun()
    else: st.sidebar.warning("Please check the confirmation box first!")

st.sidebar.write("---")
st.sidebar.subheader("📅 Manage Operational Weeks")
new_week_input = st.sidebar.text_input("Create New Week Name", placeholder="e.g. 6-10 July, 2026", key="nw_inp")
if st.sidebar.button("➕ Add Custom Week Block", use_container_width=True, key="add_wk_master_btn"):
    if new_week_input.strip():
        current_laptop_time = get_laptop_time()
        conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO operational_weeks (week_name, timestamp) VALUES (?, ?)", (new_week_input.strip(), current_laptop_time))
            conn.commit()
            st.sidebar.success("Week Added!")
            st.rerun()
        except sqlite3.IntegrityError: st.sidebar.warning("Week block already exists.")
        finally: conn.close()

st.sidebar.write("")
del_week_select = st.sidebar.selectbox("Select Specific Week Block to Delete", week_options, key="dw_sel_v20")
confirm_del_wk = st.sidebar.checkbox("Confirm Single Week Block Deletion", key="c_del_wk_v20")
if st.sidebar.button("🗑️ Delete Single Selected Week Block", use_container_width=True, type="primary", key="del_single_wk_btn"):
    if confirm_del_wk:
        conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
        cursor = conn.cursor()
        cursor.execute("DELETE FROM operational_weeks WHERE week_name = ?", (del_week_select,))
        cursor.execute("DELETE FROM trades WHERE week_block = ?", (del_week_select,))
        cursor.execute("DELETE FROM cash_transactions WHERE week_block = ?", (del_week_select,))
        conn.commit()
        conn.close()
        st.sidebar.error(f"Week '{del_week_select}' & its data wiped safely!")
        st.rerun()
    else: st.sidebar.warning("Single delete blocked! Check confirmation box first.")

st.sidebar.write("---")
confirm_purge_wk = st.sidebar.checkbox("Confirm Purging All Weeks", key="cp_wk_chk")
if st.sidebar.button("🗑️ Purge All Custom Week Blocks", use_container_width=True, type="primary", key="p_weeks_btn"):
    if confirm_purge_wk:
        conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
        cursor = conn.cursor()
        cursor.execute("DELETE FROM operational_weeks")
        conn.commit()
        conn.close()
        st.sidebar.error("All weeks cleared! Restarting fresh template.")
        st.rerun()
    else: st.sidebar.warning("Check confirmation box!")
# ==========================================
# [PART_7_END]
# ==========================================
# ==========================================
# [PART_8_START] - Pure Core Structured Ledger Data Engine
# ==========================================
def get_client_ledger_data(client_name, week_block):
    conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    cursor = conn.cursor()
    cursor.execute("SELECT opening_balance FROM client_settings WHERE client_name = ?", (client_name,))
    row = cursor.fetchone()
    opening_base = float(row[0]) if row and row[0] is not None else 0.0
    
    df_cash = pd.read_sql_query("SELECT * FROM cash_transactions WHERE client_name = ? AND week_block = ?", conn, params=(client_name, week_block))
    df_trades = pd.read_sql_query("SELECT * FROM trades WHERE client_name = ? AND week_block = ?", conn, params=(client_name, week_block))
    conn.close()
    
    total_deposit = float(df_cash[df_cash['tx_type'] == 'DEPOSIT']['amount'].sum()) if not df_cash.empty else 0.0
    total_withdraw = float(df_cash[df_cash['tx_type'] == 'WITHDRAWAL']['amount'].sum()) if not df_cash.empty else 0.0
    total_turnover = float(df_trades['turnover'].sum()) if not df_trades.empty else 0.0
# ==========================================
# [PART_8_END]
# ==========================================
# ==========================================
# [PART_9_START] - Core Calculations Framework (Live Rates Connection Fixed)
# ==========================================
    if not df_trades.empty:
        df_trades['script_name'] = df_trades['script_name'].astype(str).str.lower().str.strip()
        # 🔥 CRITICAL CASE INSENSITIVE UNIFICATION PATCH
        df_trades['selected_expiry'] = df_trades['selected_expiry'].astype(str).str.lower().str.strip()
        
        for (name, expiry), group in df_trades.groupby(['script_name', 'selected_expiry']):
            b_vol = int(group['buy_qty'].sum())
            s_vol = int(group['sell_qty'].sum())
            matched_qty = min(b_vol, s_vol)
            if matched_qty > 0:
                avg_b = (group['buy_qty'] * group['buy_price']).sum() / b_vol if b_vol > 0 else 0.0
                avg_s = (group['sell_qty'] * group['sell_price']).sum() / s_vol if s_vol > 0 else 0.0
                auto_pnl = (avg_s - avg_b) * matched_qty
                
                mask = (df_trades['script_name'] == name) & (df_trades['selected_expiry'] == expiry)
                df_trades.loc[mask, 'manual_pnl'] = auto_pnl / len(group)
                df_trades.loc[mask, 'status'] = 'SETTLED & MATCHED'

    if not df_trades.empty:
        df_trades['manual_pnl'] = pd.to_numeric(df_trades['manual_pnl'], errors='coerce').fillna(0.0)
        df_trades['brokerage'] = pd.to_numeric(df_trades['brokerage'], errors='coerce').fillna(0.0)
        total_brokerage = float(df_trades['brokerage'].sum())
        net_realized_pnl = float(df_trades['manual_pnl'].sum()) - total_brokerage
    else:
        total_brokerage, net_realized_pnl = 0.0, 0.0
    
    return {
        "opening_balance": opening_base, "total_deposit": total_deposit, "total_withdraw": total_withdraw,
        "total_turnover": total_turnover, "total_brokerage": total_brokerage, "net_realized_pnl": net_realized_pnl,
        "trades_df": df_trades, "cash_df": df_cash
    }
# ==========================================
# [PART_9_END]
# ==========================================
# ==========================================
# [PART_10_START] - Weekly Report Generator Engine (Expiry & Crash Fixed)
# ==========================================
def generate_pdf_report(client_name, week_block, metrics, live_mtm_total):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=20, leftMargin=20, topMargin=25, bottomMargin=25)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontSize=15, textColor=colors.HexColor("#0d47a1"), spaceAfter=10)
    section_style = ParagraphStyle('SecHead', parent=styles['Heading2'], fontSize=10, textColor=colors.HexColor("#1e88e5"), spaceBefore=12, spaceAfter=6)
    cell_style = ParagraphStyle('CWrap', parent=styles['Normal'], fontSize=7.5, leading=9, alignment=1)
    cell_bold = ParagraphStyle('CWrapB', parent=styles['Normal'], fontSize=7.5, leading=9, fontName="Helvetica-Bold", alignment=1)
    header_style_main = ParagraphStyle('HeadMain', parent=styles['Normal'], fontSize=7.5, leading=9, fontName="Helvetica-Bold", textColor=colors.HexColor("#1e3a8a"), alignment=1)
    
    story.append(Paragraph("Salasar Wealth - Executive Account Statement", title_style))
    story.append(Paragraph(f"<b>Client Profile:</b> {client_name} | <b>Week Accounting Block:</b> {week_block}", ParagraphStyle('MetaSub', parent=styles['Normal'], fontSize=8.5, leading=10)))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("Account Performance Summary", section_style))
    final_closing = metrics['opening_balance'] + metrics['total_deposit'] - metrics['total_withdraw'] + metrics['net_realized_pnl'] + live_mtm_total
    
    summary_data = [
        [Paragraph("Opening Base Capital Balance", ParagraphStyle('L1', parent=styles['Normal'], fontSize=7.5)), f"{metrics['opening_balance']:,.2f}"],
        [Paragraph("Current Week Deposits (+)", ParagraphStyle('L2', parent=styles['Normal'], fontSize=7.5)), f"{metrics['total_deposit']:,.2f}"],
        [Paragraph("Current Week Withdrawals (-)", ParagraphStyle('L3', parent=styles['Normal'], fontSize=7.5)), f"{metrics['total_withdraw']:,.2f}"],
        [Paragraph("Net Realized Trading P&L (Closed/Matched)", ParagraphStyle('L4', parent=styles['Normal'], fontSize=7.5)), f"{metrics['net_realized_pnl']:,.2f}"],
        [Paragraph("Live Floating MTM Box Total", ParagraphStyle('L5', parent=styles['Normal'], fontSize=7.5)), f"{live_mtm_total:,.2f}"],
        [Paragraph("<b>Final Current Balance / Net Asset Value</b>", ParagraphStyle('L6', parent=styles['Normal'], fontSize=7.5, fontName="Helvetica-Bold")), f"{final_closing:,.2f}"]
    ]
    t_sum = Table(summary_data, colWidths=None)
    t_sum.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#eaeaea")),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#e3f2fd")),
        ('PADDING', (0,0), (-1,-1), 4), ('ALIGN', (1,0), (1,-1), 'RIGHT')
    ]))
    story.append(t_sum)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Weekly Capital Movements Book (Cash Logs)", section_style))
    cash_headers = [Paragraph("Transaction ID", header_style_main), Paragraph("Type", header_style_main), Paragraph("Amount Volume (₹)", header_style_main), Paragraph("Narration / Remarks", header_style_main), Paragraph("System Log Time", header_style_main)]
    cash_rows = [cash_headers]
    
    if not metrics['cash_df'].empty:
        for _, r in metrics['cash_df'].iterrows():
            clean_ts = str(r['timestamp'])[:16] if r['timestamp'] else "-"
            cash_rows.append([str(r['id']), str(r['tx_type']).upper(), f"{r['amount']:,.2f}", str(r['remarks']), clean_ts])
        t_cash_block = Table(cash_rows, colWidths=None)
        t_cash_block.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f1f5f9")), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#eaeaea")),
            ('FONTSIZE', (0,0), (-1,-1), 7.5), ('PADDING', (0,0), (-1,-1), 4), ('ALIGN', (0,0), (-1,-1), 'CENTER')
        ]))
        story.append(t_cash_block)
    else:
        story.append(Paragraph("*No dynamic capital deposits or withdrawal entries logged.*", ParagraphStyle('N1', parent=styles['Normal'], fontSize=7.5)))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Executed Trade Ledger Book (Weekly Records)", section_style))
    headers = ["ID", "Script", "Expiry", "Type", "B_Qty", "B_Price", "S_Qty", "S_Price", "Turnover", "Brokerage", "P&L", "Status"]
    trade_headers = [Paragraph(h, header_style_main) for h in headers]
    trade_rows = [trade_headers]
    
    if not metrics['trades_df'].empty:
        for _, r in metrics['trades_df'].iterrows():
            trade_rows.append([
                str(r['id']), str(r['script_name']).upper(), str(r['selected_expiry']), str(r['action_type']), 
                str(r['buy_qty']), f"{r['buy_price']:,.2f}", str(r['sell_qty']), f"{r['sell_price']:,.2f}", 
                f"{r['turnover']:,.2f}", f"{r['brokerage']:,.2f}", f"{r['manual_pnl']:,.2f}", str(r['status'])
            ])
        t_tr = Table(trade_rows, colWidths=None)
        t_tr.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f1f5f9")), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#eaeaea")), 
            ('FONTSIZE', (0,0), (-1,-1), 6.5), ('PADDING', (0,0), (-1,-1), 3), ('ALIGN', (0,0), (-1,-1), 'CENTER')
        ]))
        story.append(t_tr)
    else:
        story.append(Paragraph("*No executed trades recorded in this week block.*", ParagraphStyle('N2', parent=styles['Normal'], fontSize=7.5)))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("Standing Net Positions Inventory Book (Carry Forward Exposure)", section_style))
    pos_rows = [[Paragraph("Scrip / Symbol Name", header_style_main), Paragraph("Expiry", header_style_main), Paragraph("Net Open Qty", header_style_main), Paragraph("Avg Cost Rate (₹)", header_style_main), Paragraph("Live Price (₹)", header_style_main), Paragraph("Floating MTM (₹)", header_style_main)]]
    
    if not metrics['trades_df'].empty:
        df_pdf_trades = metrics['trades_df'].copy()
        df_pdf_trades['script_name'] = df_pdf_trades['script_name'].astype(str).str.lower().str.strip()
        df_pdf_trades['selected_expiry'] = df_pdf_trades['selected_expiry'].astype(str).str.strip()
        
        for (name, expiry), group in df_pdf_trades.groupby(['script_name', 'selected_expiry']):
            b_v = int(group['buy_qty'].sum())
            s_v = int(group['sell_qty'].sum())
            net_q = b_v - s_v
            
            if net_q != 0: 
                avg_b = (group['buy_qty'] * group['buy_price']).sum() / b_v if b_v > 0 else 0.0
                avg_s = (group['sell_qty'] * group['sell_price']).sum() / s_v if s_v > 0 else 0.0
                avg_cost = avg_b if net_q > 0 else avg_s
                
                # 🔥 FIX BLOCK: Symmetrical query parsing to pull exact single token key rates
                live_p = get_global_price(f"{name}_{expiry}")
                if live_p <= 0:
                    live_p = get_global_price(name)
                if live_p <= 0:
                    live_p = avg_cost
                    
                calc_mtm = (live_p - avg_b) * abs(net_q) if net_q > 0 else (avg_s - live_p) * abs(net_q)
                direction = "BUY" if net_q > 0 else "SELL"
                
                pos_rows.append([
                    Paragraph(name.upper(), cell_style), Paragraph(expiry if expiry else "-", cell_style),
                    Paragraph(f"{abs(net_q)} [{direction}]", cell_style), Paragraph(f"{avg_cost:,.2f}", cell_style),
                    Paragraph(f"{live_p:,.2f}", cell_style), Paragraph(f"{calc_mtm:+,.2f}", cell_bold)
                ])
                
    if len(pos_rows) > 1:
        t_pos = Table(pos_rows, colWidths=None)
        t_pos.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#fff8e1")), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#eaeaea")), 
            ('PADDING', (0,0), (-1,-1), 5), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
        ]))
        story.append(t_pos)
    else:
        story.append(Paragraph("*No active open standing positions recorded.*", ParagraphStyle('N3', parent=styles['Normal'], fontSize=7.5)))
        
    doc.build(story)
    buffer.seek(0)
    return buffer
# ==========================================
# [PART_10_END]
# ==========================================
# ==========================================
# [PART_11_START] - Historical Statement Engine Initial Core Layer
# ==========================================
def generate_complete_historical_ledger_pdf(client_name):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=25, leftMargin=25, topMargin=30, bottomMargin=30)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('HistTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor("#111827"), spaceAfter=4)
    sub_style = ParagraphStyle('HistSub', parent=styles['Normal'], fontSize=8.5, textColor=colors.HexColor("#4B5563"), spaceAfter=15)
    section_style = ParagraphStyle('HistSec', parent=styles['Heading2'], fontSize=11, textColor=colors.HexColor("#0F766E"), spaceBefore=10, spaceAfter=8)
    cell_style = ParagraphStyle('HCell', parent=styles['Normal'], fontSize=8, leading=10)
    cell_bold = ParagraphStyle('HCellB', parent=styles['Normal'], fontSize=8, leading=10, fontName="Helvetica-Bold")
    cell_right = ParagraphStyle('HCellR', parent=styles['Normal'], fontSize=8, leading=10, alignment=2)
    cell_right_bold = ParagraphStyle('HCellRB', parent=styles['Normal'], fontSize=8, leading=10, fontName="Helvetica-Bold", alignment=2)
    
    story.append(Paragraph("SALASAR WEALTH", title_style))
    story.append(Paragraph(f"<b>Statement Of Account (Historical Ledger)</b> | Client Profile: <b>{client_name.upper()}</b> | Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M')}", sub_style))
    
    conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    cursor = conn.cursor()
    cursor.execute("SELECT opening_balance FROM client_settings WHERE client_name = ?", (client_name,))
    row = cursor.fetchone()
    base_opening = float(row[0]) if row and row is not None else 0.0
    
    df_all_cash = pd.read_sql_query("SELECT * FROM cash_transactions WHERE client_name = ?", conn, params=(client_name,))
    df_all_trades = pd.read_sql_query("SELECT * FROM trades WHERE client_name = ?", conn, params=(client_name,))
    
    cursor.execute("SELECT week_name FROM operational_weeks ORDER BY timestamp ASC")
    all_timeline_weeks = [str(r[0]) for r in cursor.fetchall()]
    conn.close()
    
    story.append(Paragraph("Chronological Transaction Log Sheet", section_style))
    headers = [
        Paragraph("<b>Date / Time</b>", cell_bold), Paragraph("<b>Transaction Narration Particulars</b>", cell_bold),
        Paragraph("<b>Debit (Dr)</b>", cell_bold), Paragraph("<b>Credit (Cr)</b>", cell_bold),
        Paragraph("<b>Running Balance</b>", cell_bold)
    ]
    ledger_rows = [headers]
    
    running_bal = base_opening
    ledger_rows.append([
        Paragraph("-", cell_style), Paragraph("<b>BASE ACCOUNT INITIAL OPENING BALANCE</b>", cell_bold),
        "-", "-", f"{running_bal:,.2f}"
    ])
# ==========================================
# [PART_11_END]
# ==========================================
# ==========================================
# [PART_12_START] - Historical Statement Engine Extended Logic Layer
# ==========================================
    for w_block in all_timeline_weeks:
        df_w_cash = df_all_cash[df_all_cash['week_block'] == w_block] if not df_all_cash.empty else pd.DataFrame()
        df_w_trades = df_all_trades[df_all_trades['week_block'] == w_block] if not df_all_trades.empty else pd.DataFrame()
        
        if not df_w_cash.empty:
            for _, r in df_w_cash.iterrows():
                remarks_str = str(r['remarks'])
                if "Weekly Settlement Closing Entry" in remarks_str or "Weekly Settlement Opening Entry" in remarks_str:
                    continue
                    
                dr_amt = float(r['amount']) if r['tx_type'] == 'WITHDRAWAL' else 0.0
                cr_amt = float(r['amount']) if r['tx_type'] == 'DEPOSIT' else 0.0
                running_bal = running_bal + cr_amt - dr_amt
                
                dr_str = f"{dr_amt:,.2f}" if dr_amt > 0 else "-"
                cr_str = f"{cr_amt:,.2f}" if cr_amt > 0 else "-"
                clean_ts = str(r['timestamp'])[:16] if r['timestamp'] else "-"
                
                ledger_rows.append([
                    Paragraph(clean_ts, cell_style),
                    Paragraph(f"Cash {r['tx_type']} [{w_block}] - {r['remarks']}", cell_style),
                    Paragraph(dr_str, cell_right), Paragraph(cr_str, cell_right),
                    Paragraph(f"{running_bal:,.2f}", cell_right_bold)
                ])

        if not df_w_trades.empty:
            working_trades = df_w_trades.copy()
            working_trades['script_name'] = working_trades['script_name'].astype(str).str.lower().str.strip()
            working_trades['selected_expiry'] = working_trades['selected_expiry'].astype(str).str.strip()
            
            for (name_scr, exp_scr), group_scr in working_trades.groupby(['script_name', 'selected_expiry']):
                b_v = int(group_scr['buy_qty'].sum())
                s_v = int(group_scr['sell_qty'].sum())
                matched_q = min(b_v, s_v)
                if matched_q > 0:
                    avg_b = (group_scr['buy_qty'] * group_scr['buy_price']).sum() / b_v if b_v > 0 else 0.0
                    avg_s = (group_scr['sell_qty'] * group_scr['sell_price']).sum() / s_v if s_v > 0 else 0.0
                    auto_p = (avg_s - avg_b) * matched_q
                    mask = (working_trades['script_name'] == name_scr) & (working_trades['selected_expiry'] == exp_scr)
                    working_trades.loc[mask, 'manual_pnl'] = auto_p / len(group_scr)

            working_trades['manual_pnl'] = pd.to_numeric(working_trades['manual_pnl'], errors='coerce').fillna(0.0)
            working_trades['brokerage'] = pd.to_numeric(working_trades['brokerage'], errors='coerce').fillna(0.0)
            
            net_block_realized_pnl_raw = float(working_trades['manual_pnl'].sum())
            total_w_brokerage = float(working_trades['brokerage'].sum())
            last_t_stamp = str(working_trades['timestamp'].max())[:16] if 'timestamp' in working_trades.columns else "-"
            
            if net_block_realized_pnl_raw != 0:
                dr_val = abs(net_block_realized_pnl_raw) if net_block_realized_pnl_raw < 0 else 0.0
                cr_val = net_block_realized_pnl_raw if net_block_realized_pnl_raw > 0 else 0.0
                running_bal = running_bal + cr_val - dr_val
                dr_str = f"{dr_val:,.2f}" if dr_val > 0 else "-"
                cr_str = f"{cr_val:,.2f}" if cr_val > 0 else "-"
                
                ledger_rows.append([
                    Paragraph(last_t_stamp, cell_style),
                    Paragraph(f"Net Realized Trading P&L Results [{w_block}]", cell_style),
                    Paragraph(dr_str, cell_right), Paragraph(cr_str, cell_right),
                    Paragraph(f"{running_bal:,.2f}", cell_right_bold)
                ])
                
            if total_w_brokerage > 0:
                running_bal -= total_w_brokerage
                ledger_rows.append([
                    Paragraph(last_t_stamp, cell_style),
                    Paragraph(f"Total Weekly Brokerage Cost Friction [{w_block}]", cell_style),
                    Paragraph(f"{total_w_brokerage:,.2f}", cell_right), Paragraph("-", cell_right),
                    Paragraph(f"{running_bal:,.2f}", cell_right_bold)
                ])
# ==========================================
# [PART_12_END]
# ==========================================
# ==========================================
# [PART_13_START] - Historical Statement Engine PDF Render Layer (Live MTM Restored)
# ==========================================
        active_week_selected_global = st.session_state.get('active_block', '')
        if w_block != active_week_selected_global:
            hist_floating_mtm_calc = 0.0
            has_active_floating_exposure = False
            
            if not df_w_trades.empty:
                df_w_trades_clean = df_w_trades.copy()
                df_w_trades_clean['script_name'] = df_w_trades_clean['script_name'].astype(str).str.lower().str.strip()
                df_w_trades_clean['selected_expiry'] = df_w_trades_clean['selected_expiry'].astype(str).str.strip()
                
                for (name_s, exp_s), group_s in df_w_trades_clean.groupby(['script_name', 'selected_expiry']):
                    b_v = int(group_s['buy_qty'].sum())
                    s_v = int(group_s['sell_qty'].sum())
                    n_stand = b_v - s_v
                    
                    if n_stand != 0:
                        has_active_floating_exposure = True
                        avg_b_p = (group_s['buy_qty'] * group_s['buy_price']).sum() / b_v if b_v > 0 else 0.0
                        avg_s_p = (group_s['sell_qty'] * group_s['sell_price']).sum() / s_v if s_v > 0 else 0.0
                        
                        conn_hist = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                        cursor_h = conn_hist.cursor()
                        cursor_h.execute("SELECT week_name FROM operational_weeks ORDER BY timestamp ASC")
                        all_weeks_list = [str(r[0]) for r in cursor_h.fetchall()]
                        
                        timeline_idx = all_weeks_list.index(w_block) if w_block in all_weeks_list else 0
                        lock_rate = 0.0
                        
                        if timeline_idx + 1 < len(all_weeks_list):
                            next_forward_week = all_weeks_list[timeline_idx + 1]
                            df_next_carry = pd.read_sql_query(
                                "SELECT buy_price, sell_price FROM trades WHERE client_name = ? AND week_block = ? AND script_name = ? AND selected_expiry = ? AND status = 'CARRY FORWARD'", 
                                conn_hist, params=(client_name, next_forward_week, name_s, exp_s)
                            )
                            if not df_next_carry.empty:
                                lock_rate = float(df_next_carry.iloc[0]['buy_price'] if n_stand > 0 else df_next_carry.iloc[0]['sell_price'])
                                
                        if lock_rate == 0.0:
                            cursor_h.execute("SELECT closing_price FROM global_market_prices WHERE symbol = ?", (f"{name_s}_{exp_s}",))
                            row_p = cursor_h.fetchone()
                            if not row_p:
                                cursor_h.execute("SELECT closing_price FROM global_market_prices WHERE symbol = ?", (name_s,))
                                row_p = cursor_h.fetchone()
                            lock_rate = float(row_p[0]) if row_p and row_p is not None else (avg_b_p if n_stand > 0 else avg_s_p)
                        conn_hist.close()
                        
                        calc_m_val = (lock_rate - avg_b_p) * abs(n_stand) if n_stand > 0 else (avg_s_p - lock_rate) * abs(n_stand)
                        hist_floating_mtm_calc += calc_m_val
                        
            if has_active_floating_exposure and hist_floating_mtm_calc != 0:
                dr_mtm = abs(hist_floating_mtm_calc) if hist_floating_mtm_calc < 0 else 0.0
                cr_mtm = hist_floating_mtm_calc if hist_floating_mtm_calc > 0 else 0.0
                running_bal = running_bal + cr_mtm - dr_mtm
                dr_mtm_str = f"{dr_mtm:,.2f}" if dr_mtm > 0 else "-"
                cr_mtm_str = f"{cr_mtm:,.2f}" if cr_mtm > 0 else "-"
                
                try: clean_ts_mtm = str(df_w_trades['timestamp'].max())[:16]
                except: clean_ts_mtm = "-"
                    
                ledger_rows.append([
                    Paragraph(clean_ts_mtm, cell_style),
                    Paragraph(f"<b>Settled Open Positions Closed Floating MTM Lock [{w_block}]</b>", cell_style),
                    Paragraph(dr_mtm_str, cell_right), Paragraph(cr_mtm_str, cell_right),
                    Paragraph(f"{running_bal:,.2f}", cell_right_bold)
                ])

    # 🔥 RESTORED ENGINE LAYER: Running/Active open standing live MTM valuation trace mapping
    live_floating_mtm_calc = 0.0
    active_week_selected = st.session_state.get('active_block', '')
    if not df_all_trades.empty and active_week_selected:
        df_active_trades = df_all_trades[df_all_trades['week_block'] == active_week_selected].copy()
        df_active_trades['script_name'] = df_active_trades['script_name'].astype(str).str.lower().str.strip()
        df_active_trades['selected_expiry'] = df_active_trades['selected_expiry'].astype(str).str.strip()
        
        for (name_s, exp_s), group_s in df_active_trades.groupby(['script_name', 'selected_expiry']):
            b_v = int(group_s['buy_qty'].sum())
            s_v = int(group_s['sell_qty'].sum())
            n_stand = b_v - s_v
            if n_stand != 0:
                avg_b_p = (group_s['buy_qty'] * group_s['buy_price']).sum() / b_v if b_v > 0 else 0.0
                avg_s_p = (group_s['sell_qty'] * group_s['sell_price']).sum() / s_v if s_v > 0 else 0.0
                
                conn_p = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                cursor_p = conn_p.cursor()
                cursor_p.execute("SELECT closing_price FROM global_market_prices WHERE symbol = ?", (f"{name_s}_{exp_s}",))
                row_p = cursor_p.fetchone()
                if not row_p:
                    cursor_p.execute("SELECT closing_price FROM global_market_prices WHERE symbol = ?", (name_s,))
                    row_p = cursor_p.fetchone()
                conn_p.close()
                g_rate = float(row_p[0]) if row_p and row_p is not None else 0.0
                
                live_in_val = float(g_rate if g_rate > 0 else (avg_b_p if n_stand > 0 else avg_s_p))
                calc_m_val = (live_in_val - avg_b_p) * abs(n_stand) if n_stand > 0 else (avg_s_p - live_in_val) * abs(n_stand)
                live_floating_mtm_calc += calc_m_val
                
    if live_floating_mtm_calc != 0:
        dr_live = abs(live_floating_mtm_calc) if live_floating_mtm_calc < 0 else 0.0
        cr_live = live_floating_mtm_calc if live_floating_mtm_calc > 0 else 0.0
        running_bal = running_bal + cr_live - dr_live
        dr_live_str = f"{dr_live:,.2f}" if dr_live > 0 else "-"
        cr_live_str = f"{cr_live:,.2f}" if cr_live > 0 else "-"
        current_time_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        ledger_rows.append([
            Paragraph(current_time_str, cell_style),
            Paragraph(f"<b>Running Active Open Positions Live Floating MTM [{active_week_selected}]</b>", cell_style),
            Paragraph(dr_live_str, cell_right), Paragraph(cr_live_str, cell_right),
            Paragraph(f"{running_bal:,.2f}", cell_right_bold)
        ])

    t_ledger = Table(ledger_rows, colWidths=None)
    t_ledger.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0f766e")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('PADDING', (0,0), (-1,-1), 4),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
    ]))
    story.append(t_ledger)
    
    doc.build(story)
    buffer.seek(0)
    return buffer
# ==========================================
# [PART_13_END]
# ==========================================
# ==========================================
# [PART_14_START] - Only Net Positions Report Engine (High Contrast View)
# ==========================================
def generate_only_net_positions_pdf(client_name, week_block, trades_df):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=35, bottomMargin=35)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('NetTitle', parent=styles['Heading1'], fontSize=15, textColor=colors.HexColor("#0d47a1"), spaceAfter=5)
    meta_style = ParagraphStyle('NetMeta', parent=styles['Normal'], fontSize=8.5, textColor=colors.HexColor("#4B5563"), spaceAfter=15)
    cell_style = ParagraphStyle('NetCell', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)
    cell_bold = ParagraphStyle('NetCellB', parent=styles['Normal'], fontSize=8, leading=10, fontName="Helvetica-Bold", alignment=1)
    
    # 🔥 HIGH CONTRAST HEADER CELL STYLE FOR WEAK EYESIGHT
    header_style = ParagraphStyle('NetHead', parent=styles['Normal'], fontSize=8.5, leading=10, fontName="Helvetica-Bold", textColor=colors.HexColor("#1e3a8a"), alignment=1)
    
    story.append(Paragraph("SALASAR WEALTH - STANDING NET POSITIONS REPORT", title_style))
    story.append(Paragraph(f"<b>Client Profile:</b> {client_name.upper()} | <b>Week Account Block:</b> {week_block} | <b>Status:</b> Live Exposure Active", meta_style))
    
    # Upgraded Matrix Columns Headers with clear styling blocks
    pos_rows = [[
        Paragraph("Scrip / Symbol Name", header_style),
        Paragraph("Net Open Qty", header_style),
        Paragraph("Avg Cost Rate (₹)", header_style),
        Paragraph("Live Price (₹)", header_style),
        Paragraph("Floating MTM (₹)", header_style)
    ]]
    
    if not trades_df.empty:
        for name, group in trades_df.groupby('script_name'):
            b_v = int(group['buy_qty'].sum())
            s_v = int(group['sell_qty'].sum())
            net_q = b_v - s_v
            
            if net_q != 0: 
                avg_b = (group['buy_qty'] * group['buy_price']).sum() / b_v if b_v > 0 else 0.0
                avg_s = (group['sell_qty'] * group['sell_price']).sum() / s_v if s_v > 0 else 0.0
                avg_cost = avg_b if net_q > 0 else avg_s
                
                import sqlite3
                conn_pdf = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                cursor_pdf = conn_pdf.cursor()
                cursor_pdf.execute("SELECT closing_price FROM global_market_prices WHERE symbol = ?", (str(name).lower().strip(),))
                row_p = cursor_pdf.fetchone()
                conn_pdf.close()
                
                live_p = float(row_p[0]) if row_p and row_p is not None else avg_cost
                calc_mtm = (live_p - avg_b) * abs(net_q) if net_q > 0 else (avg_s - live_p) * abs(net_q)
                direction = "BUY" if net_q > 0 else "SELL"
                
                pos_rows.append([
                    Paragraph(name.upper(), cell_style),
                    Paragraph(f"{abs(net_q)} [{direction}]", cell_style),
                    Paragraph(f"{avg_cost:,.2f}", cell_style),
                    Paragraph(f"{live_p:,.2f}", cell_style),
                    Paragraph(f"{calc_mtm:+,.2f}", cell_bold)
                ])
                
    if len(pos_rows) > 1:
        t_pos = Table(pos_rows, colWidths=None)
        t_pos.setStyle(TableStyle([
            # 🔥 HIGH-VISIBILITY COLOR THEME: Light pastel blue background with crystal clear borders
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#dbeafe")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#93c5fd")),
            ('PADDING', (0,0), (-1,-1), 6),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
        ]))
        story.append(t_pos)
    else:
        story.append(Paragraph("<b>* No active open standing positions recorded for this week block. *</b>", cell_style))
        
    doc.build(story)
    buffer.seek(0)
    return buffer
# ==========================================
# [PART_14_END]
# ==========================================
# ==========================================
# [PART_15_START] - Rollover Position Matrix Processor (Expiry Fixed)
# ==========================================
def execute_advanced_weekly_settlement(client, current_week, target_week):
    metrics = get_client_ledger_data(client, current_week)
    pre_live_mtm = 0.0
    conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    cursor = conn.cursor()
    
    cursor.execute("DELETE FROM trades WHERE client_name = ? AND week_block = ? AND status = 'CARRY FORWARD'", (client, target_week))
    
    if not metrics["trades_df"].empty:
        df_t = metrics["trades_df"].copy()
        for (name, expiry), group in df_t.groupby(['script_name', 'selected_expiry']):
            b_vol = int(group['buy_qty'].sum())
            s_vol = int(group['sell_qty'].sum())
            avg_buy = (group['buy_qty'] * group['buy_price']).sum() / b_vol if b_vol > 0 else 0.0
            avg_sell = (group['sell_qty'] * group['sell_price']).sum() / s_vol if s_vol > 0 else 0.0
            net_stand = b_vol - s_vol
            
            if net_stand != 0:
                global_saved_rate = get_global_price(f"{name}_{expiry}")
                if global_saved_rate <= 0:
                    global_saved_rate = get_global_price(name)
                bhav_rate = float(global_saved_rate if global_saved_rate > 0 else (avg_buy if net_stand > 0 else avg_sell))
                calc_turnover = abs(net_stand) * bhav_rate
                current_laptop_time = get_laptop_time()
                
                # 🔥 CRITICAL FIX: Inserting 'expiry' variable directly into database instead of empty '' string
                if net_stand > 0:
                    cursor.execute("""
                        INSERT INTO trades (client_name, week_block, exchange, script_name, selected_expiry, action_type, buy_qty, buy_price, sell_qty, sell_price, turnover, brokerage, manual_pnl, status, timestamp) 
                        VALUES (?, ?, 'NSE', ?, ?, 'BUY', ?, ?, 0, 0.0, ?, 0.0, 0, 'CARRY FORWARD', ?)
                    """, (client, target_week, name, expiry, net_stand, bhav_rate, calc_turnover, current_laptop_time))
                else:
                    cursor.execute("""
                        INSERT INTO trades (client_name, week_block, exchange, script_name, selected_expiry, action_type, buy_qty, buy_price, sell_qty, sell_price, turnover, brokerage, manual_pnl, status, timestamp) 
                        VALUES (?, ?, 'NSE', ?, ?, 'SELL', 0, 0.0, ?, ?, ?, 0.0, 0, 'CARRY FORWARD', ?)
                    """, (client, target_week, name, expiry, abs(net_stand), bhav_rate, calc_turnover, current_laptop_time))
                    
                calc_mtm = (bhav_rate - avg_buy) * abs(net_stand) if net_stand > 0 else (avg_sell - bhav_rate) * abs(net_stand)
                pre_live_mtm += calc_mtm

    net_cash_flow = float(metrics["total_deposit"]) - float(metrics["total_withdraw"])
    screen_final_balance = float(metrics["opening_balance"]) + net_cash_flow + float(metrics["net_realized_pnl"]) + pre_live_mtm
    offsetting_volume = -1.0 * screen_final_balance
    current_laptop_time = get_laptop_time()
    
    cursor.execute("DELETE FROM cash_transactions WHERE client_name = ? AND week_block = ? AND remarks = 'Weekly Settlement Closing Entry'", (client, current_week))
    cursor.execute("DELETE FROM cash_transactions WHERE client_name = ? AND week_block = ? AND remarks = 'Weekly Settlement Opening Entry'", (client, target_week))
    
    if offsetting_volume != 0:
        type_current_week = "DEPOSIT" if offsetting_volume >= 0 else "WITHDRAWAL"
        type_target_week = "WITHDRAWAL" if offsetting_volume >= 0 else "DEPOSIT"
        absolute_offset_amt = abs(offsetting_volume)
        
        cursor.execute("INSERT INTO cash_transactions (client_name, week_block, tx_type, amount, remarks, timestamp) VALUES (?, ?, ?, ?, 'Weekly Settlement Closing Entry', ?)", (client, current_week, type_current_week, absolute_offset_amt, current_laptop_time))
        cursor.execute("INSERT INTO cash_transactions (client_name, week_block, tx_type, amount, remarks, timestamp) VALUES (?, ?, ?, ?, 'Weekly Settlement Opening Entry', ?)", (client, target_week, type_target_week, absolute_offset_amt, current_laptop_time))
        
    conn.commit()
    conn.close()
# ==========================================
# [PART_15_END]
# ==========================================
# ==========================================
# [PART_16_START] - Excel Safe Mode Parser
# ==========================================
st.markdown("#### 📂 Bulk Sync Execution Panel")
uploaded_excel_report = st.file_uploader(
    "Upload Excel Report", type=["xls", "xlsx"], 
    key="trading_app_excel_sync_uploader"
)

if uploaded_excel_report is not None:
    if st.button("🚀 Scan Excel Report & Sync Direct", use_container_width=True):
        try:
            import openpyxl
            import sqlite3
            from datetime import datetime as dt_root # Safe root alias injection
            inserted_rows_count = 0
            skipped_rows_count = 0 # Explicit fallback baseline setup
            
            wb_memory = openpyxl.load_workbook(uploaded_excel_report, data_only=True)
            ws_sheet = wb_memory.active
            
            for row_num in range(2, ws_sheet.max_row + 1):
                try:
                    raw_time = str(ws_sheet.cell(row=row_num, column=1).value or "").strip()
                    raw_exch = str(ws_sheet.cell(row=row_num, column=4).value or "NSE").strip().upper()
                    raw_scrip = str(ws_sheet.cell(row=row_num, column=5).value or "nifty").strip().lower()
                    
                    # --- RESTORED ORIGINAL STABLE DATE PARSING LOGIC MATRIX ---
                    cell_expiry_obj = ws_sheet.cell(row=row_num, column=6).value
                    raw_expiry = "30-Jun-2026"
                    
                    if cell_expiry_obj is not None:
                        clean_exp_str = str(cell_expiry_obj).strip()
                        # Checks if the string layout matches backend standard date objects format
                        if "-" in clean_exp_str and len(clean_exp_str) >= 10:
                            try:
                                # Extracts pure date slice block before hours/minutes sequence splits
                                pure_date_token = clean_exp_str.split(" ")[0]
                                parsed_dt = dt_root.strptime(pure_date_token, "%Y-%m-%d")
                                raw_expiry = parsed_dt.strftime("%d-%b-%Y")
                            except Exception:
                                raw_expiry = clean_exp_str
                        else:
                            raw_expiry = clean_exp_str
                                
                    raw_type = str(ws_sheet.cell(row=row_num, column=7).value or "BUY").strip().upper()
# ==========================================
# [PART_16_END]
# ==========================================
# ==========================================
# [PART_17_START] - Numeric Multipliers & Anti-Duplicate Collision Lock Engine
# ==========================================
                    val_qty_raw = str(ws_sheet.cell(row=row_num, column=8).value or "").replace(',', '').strip()
                    val_prc_raw = str(ws_sheet.cell(row=row_num, column=9).value or "").replace(',', '').strip()
                    
                    if not val_qty_raw or not val_prc_raw or val_qty_raw == "" or val_prc_raw == "":
                        continue
                        
                    trade_volume_qty = int(float(val_qty_raw))
                    trade_execution_price = float(val_prc_raw)
                    
                    if trade_volume_qty <= 0 or trade_execution_price <= 0:
                        continue
                        
                    trade_direction = "SELL" if "SELL" in raw_type or "S" == raw_type else "BUY"
                    current_exchange_header = "MCX" if "MCX" in raw_exch else ("GIFTY" if "GIFT" in raw_exch or "SGX" in raw_exch else "NSE")
                    current_scrip_header = raw_scrip if raw_scrip and raw_scrip != "nan" else "nifty"
                    current_expiry_header = raw_expiry if raw_expiry and raw_expiry != "nan" else "30-Jun-2026"
                    
                    # 🔥 AUTOMATED COMMODITY LOT SIZE MULTIPLIER MATRIX ENGINE (UPGRADED EXACT COLLISION SHIELD)
                    if current_exchange_header == "MCX":
                        scrip_clean_lower = str(current_scrip_header).lower().strip()
                        
                        # Dynamic explicit commodity lot-size dictionary array mapping
                        commodity_multiplier_map = {
                            "natgasmini": 250,
                            "naturalgasm": 250,
                            "naturalgas": 1250,
                            "copper": 2500,
                            "copperm": 2500,
                            "leadmini": 1000,
                            "leadm": 1000,
                            "zincmini": 1000,
                            "zincm": 1000,
                            "zinc": 5000,
                            "lead": 5000,
                            "silvermic": 1,
                            "silverm": 5,
                            "goldm": 10,
                            "crudeoilm": 10,
                            "crudeoil": 100
                        }
                        
                        # 🔥 SOLID LONGEST-MATCH SORTING LOGIC: Fixes the dictionary looping crash completely
                        matched_multiplier_factor = None
                        
                        # Keys ko length ke hisab se descending order me sort karenge (silvermic pehle aayega, silverm baad me)
                        sorted_mcx_keys = sorted(list(commodity_multiplier_map.keys()), key=len, reverse=True)
                        
                        for mcx_key in sorted_mcx_keys:
                            if mcx_key in scrip_clean_lower:
                                matched_multiplier_factor = commodity_multiplier_map[mcx_key]
                                break
                                
                        if matched_multiplier_factor is not None:
                            # Lots ko direct contract volume size se auto-multiply karein safely
                            trade_volume_qty = int(trade_volume_qty * matched_multiplier_factor)
                            
                    calculated_turnover = float(trade_volume_qty * trade_execution_price)
                    actual_captured_timestamp = raw_time if raw_time and raw_time != "" else get_laptop_time()
                    
                    # --- CRASH PROOF BROKERAGE COMPUTATION ---
                    try:
                        conn_b = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                        cursor_b = conn_b.cursor()
                        cursor_b.execute("SELECT brokerage_type, brokerage_nse, brokerage_mcx, brokerage_gifty, per_lot_rate FROM client_settings WHERE client_name = ?", (active_client,))
                        c_sett = cursor_b.fetchone()
                        conn_b.close()
                        
                        if c_sett:
                            b_type, b_nse, b_mcx, b_gifty, l_rate = c_sett
                            if b_type == "Per Lot":
                                # Fixed standard lot rate calculation tracking using original excel raw lots
                                trade_brokerage_cost = float(int(float(val_qty_raw)) * l_rate)
                            else:
                                active_rate = b_nse if str(current_exchange_header).upper() == 'NSE' else (b_mcx if str(current_exchange_header).upper() == 'MCX' else b_gifty)
                                trade_brokerage_cost = float((calculated_turnover * active_rate) / 10000000)
                        else:
                            trade_brokerage_cost = float((calculated_turnover * 1000) / 10000000)
                    except Exception:
                        trade_brokerage_cost = float((calculated_turnover * 1000) / 10000000)

                    # --- BULLETPROOF DATE-TIME DUP MATRIX VERIFICATION MATRIX ---
                    conn_check = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                    cursor_check = conn_check.cursor()
                    cursor_check.execute("""
                        SELECT COUNT(*) FROM trades 
                        WHERE client_name = ? AND script_name = ? AND selected_expiry = ?
                        AND action_type = ? AND timestamp = ? AND (buy_price = ? OR sell_price = ?)
                    """, (active_client, current_scrip_header, current_expiry_header, trade_direction, actual_captured_timestamp, trade_execution_price, trade_execution_price))
                    exists_already = cursor_check.fetchone()
                    conn_check.close()
                    
                    if exists_already and exists_already[0] > 0:
                        skipped_rows_count += 1
                        continue
# ==========================================
# [PART_17_END]
# ==========================================
# ==========================================
# [PART_18_START] - Live Database Injection Layer (Duplicate Skip Prompts Fixed)
# ==========================================
                    conn_sync = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                    cursor_sync = conn_sync.cursor()
                    if trade_direction == "BUY":
                        cursor_sync.execute("""
                            INSERT INTO trades (client_name, week_block, exchange, script_name, selected_expiry, action_type, buy_qty, buy_price, sell_qty, sell_price, turnover, brokerage, manual_pnl, status, timestamp)
                            VALUES (?, ?, ?, ?, ?, 'BUY', ?, ?, 0, 0.0, ?, ?, 0.0, 'CARRY FORWARD', ?)
                        """, (active_client, st.session_state['active_block'], current_exchange_header, current_scrip_header, current_expiry_header, trade_volume_qty, trade_execution_price, calculated_turnover, trade_brokerage_cost, actual_captured_timestamp))
                    else:
                        cursor_sync.execute("""
                            INSERT INTO trades (client_name, week_block, exchange, script_name, selected_expiry, action_type, buy_qty, buy_price, sell_qty, sell_price, turnover, brokerage, manual_pnl, status, timestamp)
                            VALUES (?, ?, ?, ?, ?, 'SELL', 0, 0.0, ?, ?, ?, ?, 0.0, 'CARRY FORWARD', ?)
                        """, (active_client, st.session_state['active_block'], current_exchange_header, current_scrip_header, current_expiry_header, trade_volume_qty, trade_execution_price, calculated_turnover, trade_brokerage_cost, actual_captured_timestamp))
                    conn_sync.commit()
                    conn_sync.close()
                    inserted_rows_count += 1
                except Exception:
                    continue
                    
            # Safe Fallback to initialize variables if they get flushed from earlier runtime scope blocks
            if 'skipped_rows_count' not in locals(): skipped_rows_count = 0
            if 'inserted_rows_count' not in locals(): inserted_rows_count = 0
                    
            if inserted_rows_count > 0:
                st.success(f"✨ Perfect bhai! Safe Mode Engine ne aapki Excel sheet sync kar di hai. Total {inserted_rows_count} naye trades safely register kiye gaye, aur {skipped_rows_count} duplicates block skip ho gaye.")
                st.rerun()
            elif skipped_rows_count > 0:
                st.info(f"ℹ️ Is Excel file ke saare {skipped_rows_count} records exact Date-Time matching se pehle se database cockpit me saved hain! No double entries written.")
            else:
                st.info("ℹ️ Data process ho chuka hai.")
        except Exception as err: 
            st.error(f"Error executing scan: {str(err)}")

st.write("---")
# ==========================================
# [PART_18_END]
# ==========================================
# ==========================================
# [PART_19_START] - Smart Calendar Sync Engine & Layout Setup
# ==========================================
if 'active_block' not in st.session_state or st.session_state.get('global_week_selector_fixed') is None:
    default_week_idx = 0
    try:
        # Laptop ki live exact exact date pick karein
        current_date_obj = datetime.strptime(get_laptop_time(), "%Y-%m-%d %H:%M:%S").date()
        current_year_str = str(current_date_obj.year)
        
        # Months dictionary lookup to convert short name labels
        months_lookup_dict = {
            "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, 
            "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12
        }
        
        # Operational weeks array matrix ko loop chala kar check karein
        for idx, week_str in enumerate(week_options):
            if current_year_str in week_str:
                clean_week_line = str(week_str).lower().strip()
                
                # Slicing pattern to extract month names dynamically (e.g., "15-19 june, 2026" or "6-10 july, 2026")
                detected_month_num = 6 # Default fallback to June
                for month_key, month_val in months_lookup_dict.items():
                    if month_key in clean_week_line:
                        detected_month_num = month_val
                        break
                
                # Raw numbers numbers extraction sequence
                tokens_split = clean_week_line.split(",")
                date_range_part = tokens_split[0].strip() # e.g. "15-19 june" or "29 june-3 july" or "6-10 july"
                
                # Dash line matrix pattern identification
                if "-" in date_range_part:
                    dash_tokens = date_range_part.split("-")
                    start_day_digit = int("".join([c for c in dash_tokens[0] if c.isdigit()]))
                    
                    # Split further to check cross-month boundary loops (e.g. "29 june-3 july")
                    end_part_raw = dash_tokens[1].strip()
                    end_day_digit = int("".join([c for c in end_part_raw if c.isdigit()]))
                    
                    # Create perfect datetime object comparison boundaries
                    week_start_boundary = datetime(current_date_obj.year, detected_month_num, start_day_digit).date()
                    
                    # Safe handling wrapper for cross-month boundary constraints
                    if "july" in end_part_raw and "june" in clean_week_line:
                        week_end_boundary = datetime(current_date_obj.year, 7, end_day_digit).date() + timedelta(days=2)
                    else:
                        week_end_boundary = datetime(current_date_obj.year, detected_month_num, end_day_digit).date() + timedelta(days=2)
                    
                    # 🔥 LIVE SYNCHRONIZATION DETECTOR MATRIX: Lock matching row index immediately
                    if week_start_boundary - timedelta(days=2) <= current_date_obj <= week_end_boundary:
                        default_week_idx = idx
                        break
    except Exception:
        default_week_idx = 0
        
    st.session_state['active_block'] = week_options[default_week_idx] if default_week_idx < len(week_options) else week_options[0]
else:
    default_week_idx = week_options.index(st.session_state['active_block']) if st.session_state['active_block'] in week_options else 0

st.session_state['active_block'] = st.selectbox(
    "📅 Select Active Operational Week Block", week_options, index=default_week_idx, key="global_week_selector_fixed"
)

# 🔥 MAIN INTERFACE TABS DEFINITION CONFIGURATION
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["⚡ Post New Trades", "💵 Cash Ledger Entries", "📈 Account Cockpit & Reporting", "🌐 Global Price Sync Controller", "📊 Master Enterprise Dashboard", "🔍 Scrip-Wise Excel Filter"])

with tab1:
    st.subheader("📝 Searchable Order Entry Form (Fast Entry Memory Lock Active)")
    if 't_cl_fixed_matrix_key' not in st.session_state: st.session_state['t_cl_fixed_matrix_key'] = active_client
    if 'last_selected_exchange' not in st.session_state: st.session_state['last_selected_exchange'] = EXCHANGES[0] if EXCHANGES else "NSE"
    if 'last_entered_script' not in st.session_state: st.session_state['last_entered_script'] = "nifty"
    if 'last_selected_expiry' not in st.session_state: st.session_state['last_selected_expiry'] = expiry_options[0] if expiry_options else ""
    if 'last_selected_action' not in st.session_state: st.session_state['last_selected_action'] = "BUY"

    def sync_form_to_sidebar(): st.session_state['sb_cl_selector_final_matrix'] = st.session_state['t_cl_fixed_matrix_key']

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        t_client = st.selectbox("Trading Client", CLIENTS, key="t_cl_fixed_matrix_key", on_change=sync_form_to_sidebar)
        idx_ex = EXCHANGES.index(st.session_state['last_selected_exchange']) if st.session_state['last_selected_exchange'] in EXCHANGES else 0
        t_exchange = st.selectbox("Exchange Matrix", EXCHANGES, index=idx_ex, key="t_ex_fixed_matrix")
        t_manual_date = st.date_input("Execution Date", value=datetime.now().date(), key="t_manual_date_picker_block")
# ==========================================
# [PART_19_END]
# ==========================================
# ==========================================
# [PART_20_START] - Master Autocomplete Index & Order Booking Execute
# ==========================================
    with col2:
        # 🔥 STEP 1: DYNAMIC DATABASE SCRIP AUTO-DISCOVERY MATRIX
        discovered_scrips_list = []
        try:
            conn_disc = sqlite3.connect('salasar_wealth_v19_ultimate.db')
            cursor_disc = conn_disc.cursor()
            cursor_disc.execute("SELECT DISTINCT script_name FROM trades WHERE script_name IS NOT NULL AND script_name != ''")
            discovered_scrips_list = [str(r[0]).lower().strip() for r in cursor_disc.fetchall()]
            conn_disc.close()
        except Exception:
            discovered_scrips_list = []

        # 🔥 STEP 2: FULL 250+ EXPANDED HEAVY-WEIGHTS INDEX POOL DATA LIST
        built_in_scrip_pool = [
            "nifty", "banknifty", "finnifty", "midcpnifty", "reliance", "tcs", "infy", "hdfcbank", "icicibank",
            "silverm", "goldm", "crudeoil", "crudeoilm", "jswenergy", "ultracemco", "silvermic", "grasim", "jindalstel",
            "titan", "tata steel", "tatamotors", "sbi", "axisbank", "bhartiairtel", "itc", "lt", "maruti", "wipro", 
            "hcltech", "ntpc", "adannient", "adaniports", "asianpaint", "bajfinance", "bajajfinsv", "bpcl", "cipla", 
            "coalindia", "divislab", "drreddy", "eichermot", "heromotoco", "hindalco", "hindunilvr", "indusindbk", 
            "jswsteel", "kotakbank", "m&m", "nestleind", "ongc", "powergrid", "sunpharma", "techm", "tataconsum", 
            "shreecem", "upl", "vedl", "zomato", "copper", "copperm", "zinc", "zincm", "lead", "leadm", "aluminium", 
            "aluminiumm", "naturalgas", "naturalgasm", "gold", "silver", "abb", "abcafg", "abcapital", "abfrl", 
            "acc", "adanigreen", "adanipower", "alkem", "ambujacem", "apollohosp", "apollotyre", "ashokley", "astral", 
            "atgl", "auropharma", "auproperty", "bandhanbnk", "bankbaroda", "bankindia", "bataindia", 
            "bel", "bergerpaint", "bhel", "biocon", "boschltd", "canbk", "canfinhome", "chamblfert", "cholamand", 
            "coforge", "colpal", "concor", "coromandel", "cromp", "cub", "cumminsind", "dabur", "dalbhariat", 
            "deepakntr", "deltacorp", "exideind", "federalbnk", "fortis", "gail", "glenmark", "gmrinfra", "gnfc", 
            "godrejcp", "godrejprop", "granules", "gujgasltd", "gmdcltd", "hal", "havells", "hindcopper", "hindpetro", 
            "ibulhsgfin", "idbi", "idea", "idfc", "idfcfirstb", "iex", "igl", "indhotelltd", 
            "indianb", "indigo", "indusind", "iob", "irctc", "ireda", "irfc", "jindalshaw", "jkcement", "jswinfra", 
            "jublfood", "kalyanknit", "kfintech", "kotak", "liccom", "lichsgfin", "lupin", "mcdowell-n", 
            "manappuram", "maxhealth", "metropolis", "mgl", "mrf", "mcx", "muthootfin", "nationalum", "naukri", 
            "navinfluor", "nbcc", "ncc", "ofss", "pageind", "pel", "persistent", "petronet", "pfc", "pidilitind", 
            "pnb", "polycab", "poonawalla", "pvr", "radico", "rvnl", "rblbank", "sail", "shriramfin", "siemens", 
            "suntv", "syngene", "tataelxsi", "tatachg", "tatacomm", "tatapower", "torrentph", "trent", "tvsmotor", 
            "ubl", "unionbank", "voltas", "whirlpool"
        ]

        # Consolidated unique dropdown suggested entries list
        combined_suggestions = sorted(list(set(discovered_scrips_list + built_in_scrip_pool)))
        
        # 🔥 Injection of the absolute custom escape handler at index 0
        dropdown_options_final = ["➕ ENTER CUSTOM SCRIP NAME"] + combined_suggestions
        
        last_saved_scrip_val = str(st.session_state.get('last_entered_script', 'nifty')).lower().strip()
        default_dropdown_index = 0
        if last_saved_scrip_val in combined_suggestions:
            default_dropdown_index = dropdown_options_final.index(last_saved_scrip_val)
            
        selected_dropdown_token = st.selectbox(
            "Script Selection List", 
            options=dropdown_options_final, 
            index=default_dropdown_index, 
            key="t_sc_search_dropdown_box_v20"
        )
        
        # 🔥 Dynamic visibility toggle: If custom keyword is clicked, render raw text input below it
        if selected_dropdown_token == "➕ ENTER CUSTOM SCRIP NAME":
            t_script = st.text_input("Type Custom Scrip Name Manually", value="", placeholder="e.g. titan", key="t_sc_raw_manual_typed_input").lower().strip()
        else:
            t_script = selected_dropdown_token.lower().strip()
        
        idx_exp = expiry_options.index(st.session_state['last_selected_expiry']) if st.session_state['last_selected_expiry'] in expiry_options else 0
        t_expiry = st.selectbox("Expiry Schedule", expiry_options, index=idx_exp, key="t_exp_fixed_matrix")
    with col3:
        idx_act = ["BUY", "SELL"].index(st.session_state['last_selected_action'])
        t_action = st.selectbox("Action Direction", ["BUY", "SELL"], index=idx_act, key="t_act_fixed_matrix")
        t_qty = st.number_input("Quantity", min_value=0, step=1, key="t_q")
    with col4:
        t_price = st.number_input("Execution Rate (Price)", min_value=0.0, step=0.05, key="t_pr")
        apply_zero_brokerage = st.checkbox("Apply Zero Brokerage for this Trade", value=False, key="t_zero_b_chk")
        
    st.write("")
    if st.button("⚡ Execute Trade Order Booking", use_container_width=True, type="primary", key="execute_trade_matrix_btn"):
        if t_qty > 0 and t_price > 0 and t_script != "":
            st.session_state['last_selected_exchange'] = t_exchange
            st.session_state['last_entered_script'] = t_script
            st.session_state['last_selected_expiry'] = t_expiry
            st.session_state['last_selected_action'] = t_action
            
            conn_b = sqlite3.connect('salasar_wealth_v19_ultimate.db')
            cursor_b = conn_b.cursor()
            cursor_b.execute("SELECT brokerage_type, brokerage_nse, brokerage_mcx, brokerage_gifty, per_lot_rate FROM client_settings WHERE client_name = ?", (t_client,))
            c_sett = cursor_b.fetchone()
            conn_b.close()
            
            b_type, b_nse, b_mcx, b_gifty, l_rate = c_sett if c_sett else ("Per Crore", 1000, 1000, 1000, 0.0)
            calc_turnover = t_qty * t_price
            calc_brokerage = 0.0 if apply_zero_brokerage else (t_qty * l_rate if b_type == "Per Lot" else (calc_turnover * (b_nse if str(t_exchange).upper() == 'NSE' else (b_mcx if str(t_exchange).upper() == 'MCX' else b_gifty))) / 10000000)
            
            live_laptop_clock_time = datetime.now().strftime("%H:%M:%S")
            custom_targeted_timestamp = f"{t_manual_date.strftime('%d-%m-%Y')} {live_laptop_clock_time}"
            
            conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
            cursor = conn.cursor()
            if t_action == "BUY":
                cursor.execute("INSERT INTO trades (client_name, week_block, exchange, script_name, selected_expiry, action_type, buy_qty, buy_price, sell_qty, sell_price, turnover, brokerage, manual_pnl, status, timestamp) VALUES (?, ?, ?, ?, ?, 'BUY', ?, ?, 0, 0.0, ?, ?, 0.0, 'CARRY FORWARD', ?)", (t_client, st.session_state['active_block'], t_exchange, t_script, t_expiry, t_qty, t_price, calc_turnover, calc_brokerage, custom_targeted_timestamp))
            else:
                cursor.execute("INSERT INTO trades (client_name, week_block, exchange, script_name, selected_expiry, action_type, buy_qty, buy_price, sell_qty, sell_price, turnover, brokerage, manual_pnl, status, timestamp) VALUES (?, ?, ?, ?, ?, 'SELL', 0, 0.0, ?, ?, ?, ?, 0.0, 'CARRY FORWARD', ?)", (t_client, st.session_state['active_block'], t_exchange, t_script, t_expiry, t_qty, t_price, calc_turnover, calc_brokerage, custom_targeted_timestamp))
            conn.commit()
            conn.close()
            st.success("📝 Trade entry successfully written into ledger data tables!")
            st.rerun()
        else:
            st.warning("Please specify correct parameters before entry registration.")
# ==========================================
# [PART_20_END]
# ==========================================
# ==========================================
# [PART_21_START] - Capital Movements & Cash Entries Modification
# ==========================================
with tab2:
    st.subheader("💸 Capital Movements Registry (Cash Deposit / Withdrawal)")
    with st.form("cash_form", clear_on_submit=True):
        c_col1, c_col2, c_col3 = st.columns(3)
        with c_col1:
            c_client = st.selectbox("Target Client Profile", CLIENTS, key="c_cl_fixed_matrix")
            c_type = st.selectbox("Transaction Profile", ["DEPOSIT", "WITHDRAWAL"], key="c_typ_fixed_matrix")
            
            # 🔥 NEW ENTRY FIELD: Manual calendar select wrapper for targeting specific cash log dates
            c_manual_date = st.date_input("Transaction Date", value=datetime.now().date(), key="c_manual_date_picker_block")
            
        with c_col2: 
            c_amount = st.number_input("Transaction Volume (₹)", min_value=0.0, step=500.0, key="c_am")
        with c_col3: 
            c_remarks = st.text_input("Operational Remarks / Narration", value="", placeholder="Enter remarks", key="c_rem")
            
        if st.form_submit_button("💵 Process Ledger Balance Adjustment", use_container_width=True):
            # 🔥 LIVE TIME MERGE PATTERN: Takes selected calendar date and appends live clock hours/minutes/seconds
            live_laptop_clock_time = datetime.now().strftime("%H:%M:%S")
            custom_cash_timestamp = f"{c_manual_date.strftime('%Y-%m-%d')} {live_laptop_clock_time}"
            
            conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
            cursor = conn.cursor()
            cursor.execute("INSERT INTO cash_transactions (client_name, week_block, tx_type, amount, remarks, timestamp) VALUES (?, ?, ?, ?, ?, ?)", (c_client, st.session_state['active_block'], c_type, c_amount, c_remarks, custom_cash_timestamp))
            conn.commit()
            conn.close()
            st.success("💰 Cash transaction committed safely!")
            st.rerun()
            
    st.write("---")
    with st.expander("💵 Advanced Cash Entries Modification & Delete Panel"):
        t2_metrics = get_client_ledger_data(active_client, st.session_state['active_block'])
        st.subheader("📋 Active Client Cash Logs")
        if t2_metrics and not t2_metrics["cash_df"].empty:
            st.dataframe(t2_metrics["cash_df"][['id', 'client_name', 'tx_type', 'amount', 'remarks', 'timestamp']], use_container_width=True, hide_index=True)
        else:
            st.info("No cash transactions recorded.")
            
        cash_id = st.number_input("Enter Cash Transaction ID to Edit/Delete", min_value=1, step=1, key="cash_mod_id")
        if cash_id:
            conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
            df_cash_curr = pd.read_sql_query("SELECT * FROM cash_transactions WHERE id = ?", conn, params=(cash_id,))
            conn.close()
            if not df_cash_curr.empty:
                with st.form(f"cash_form_mod_{cash_id}"):
                    c_type_mod = st.selectbox("New Transaction Type", ["DEPOSIT", "WITHDRAWAL"], index=0 if df_cash_curr.loc[0, 'tx_type'] == "DEPOSIT" else 1, key=f"ctm_{cash_id}")
                    c_amt_mod = st.number_input("New Amount (₹)", value=float(df_cash_curr.loc[0, 'amount']))
                    c_rem_mod = st.text_input("New Remarks", value=str(df_cash_curr.loc[0, 'remarks']))
                    
                    bc1, bc2 = st.columns(2)
                    with bc1:
                        if st.form_submit_button("💾 Update Cash Record", use_container_width=True):
                            current_laptop_time = get_laptop_time()
                            conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                            cursor = conn.cursor()
                            cursor.execute("UPDATE cash_transactions SET tx_type=?, amount=?, remarks=?, timestamp=? WHERE id=?", (c_type_mod, c_amt_mod, c_rem_mod, current_laptop_time, cash_id))
                            conn.commit(); conn.close(); st.success("Cash entry updated!"); st.rerun()
                    with bc2:
                        chk_del_cash = st.checkbox("Confirm Cash Entry Deletion (Safety Locked)", key=f"cdc_{cash_id}")
                        if st.form_submit_button("🗑️ Delete Cash Record", use_container_width=True, type="primary"):
                            if chk_del_cash:
                                conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                                cursor = conn.cursor()
                                cursor.execute("DELETE FROM cash_transactions WHERE id = ?", (cash_id,))
                                conn.commit(); conn.close(); st.error("Cash record deleted!"); st.rerun()
                            else: st.warning("⚠️ Action blocked! Please check the safety confirmation box first.")

    # -------------------------------------------------------------------------
    # 🔥 NAYA UPGRADE: ALL CLIENTS WEEKLY CASH LOGS MATRIX VIEWER
    # -------------------------------------------------------------------------
    st.write("---")
    st.markdown("#### 📂 Master Multi-Client Weekly Cash Log Sheet")
    
    show_all_cash_matrix = st.checkbox(
        f"🔍 Show All Clients Cash Entries for [ {st.session_state['active_block']} ]", 
        value=False, 
        key="master_all_cash_sheet_view_chk"
    )
    
    if show_all_cash_matrix:
        conn_all_c = sqlite3.connect('salasar_wealth_v19_ultimate.db')
        df_all_cash_block = pd.read_sql_query(
            "SELECT id, client_name, tx_type, amount, remarks, timestamp FROM cash_transactions WHERE week_block = ? ORDER BY timestamp DESC", 
            conn_all_c, 
            params=(st.session_state['active_block'],)
        )
        conn_all_c.close()
        
        if not df_all_cash_block.empty:
            view_all_cash_df = df_all_cash_block.copy()
            view_all_cash_df.columns = ['ID', 'Client Name', 'Transaction Profile', 'Volume Amount (₹)', 'Narration Remarks', 'System Timestamp']
            
            st.dataframe(
                view_all_cash_df.style.format({
                    "Volume Amount (₹)": "{:,.2f}"
                }),
                use_container_width=True, 
                hide_index=True
            )
            
            # Aggregate calculations layer for multi-client transparency metrics
            total_block_dep = float(df_all_cash_block[df_all_cash_block['tx_type'] == 'DEPOSIT']['amount'].sum())
            total_block_wd = float(df_all_cash_block[df_all_cash_block['tx_type'] == 'WITHDRAWAL']['amount'].sum())
            net_block_flow = total_block_dep - total_block_wd
            
            st.info(f"📊 **Total Combined Week Capital Adjustments:** Total Deposits: ₹ {total_block_dep:,.2f} | Total Withdrawals: ₹ {total_block_wd:,.2f} (Net Flow: ₹ {net_block_flow:+,.2f})")
        else:
            st.info(f"Is active week block [ {st.session_state['active_block']} ] mein filhal kisi bhi client ki koi cash ledger entry saved nahi hai.")
# ==========================================
# [PART_21_END]
# ==========================================

# ==========================================
# [PART_22_START] - Cockpit Reporting Framework (Live MTM Real-Time Unfreeze)
# ==========================================
with tab3:
    st.subheader("📊 Account Cockpit Reporting")
    
    c_set_1, c_set_2 = st.columns(2)
    with c_set_2:
        next_weeks_avail = [w for w in week_options if week_options.index(w) > week_options.index(st.session_state['active_block'])]
        target_settle_week = st.selectbox("🎯 Select Next Target Week To Carry Positions", next_weeks_avail if next_weeks_avail else ["No Future Weeks Created"], key="tgt_settle_wk_drop")
        if st.button("🔄 Execute Advanced Weekly Settlement & Carry Forward", use_container_width=True, type="secondary", key="top_weekly_settle_btn"):
            if target_settle_week == "No Future Weeks Created":
                st.error("❌ Settlement failed. Sidebar me forward week block banayein!")
            else: execute_advanced_weekly_settlement(active_client, st.session_state['active_block'], target_settle_week)

    metrics = get_client_ledger_data(active_client, st.session_state['active_block'])
    pre_live_mtm_total = 0.0
    
    if metrics and not metrics["trades_df"].empty:
        df_active_calc = metrics["trades_df"][metrics["trades_df"]['week_block'] == st.session_state['active_block']].copy()
        df_active_calc['script_name'] = df_active_calc['script_name'].astype(str).str.lower().str.strip()
        df_active_calc['selected_expiry'] = df_active_calc['selected_expiry'].astype(str).str.strip()
        
        for (name, expiry), group in df_active_calc.groupby(['script_name', 'selected_expiry']):
            b_vol = int(group['buy_qty'].sum())
            s_vol = int(group['sell_qty'].sum())
            net_stand = b_vol - s_vol
            
            if net_stand != 0:
                avg_buy_price = (group['buy_qty'] * group['buy_price']).sum() / b_vol if b_vol > 0 else 0.0
                avg_sell_price = (group['sell_qty'] * group['sell_price']).sum() / s_vol if s_vol > 0 else 0.0
                
                # 🔥 REAL-TIME UNFREEZE FIX: Matching unique compound token key combination parameters
                global_price_rate = get_global_price(f"{name}_{expiry}")
                if global_price_rate <= 0:
                    global_price_rate = get_global_price(name)
                    
                live_input_val = float(global_price_rate if global_price_rate > 0 else (avg_buy_price if net_stand > 0 else avg_sell_price))
                calc_mtm = (live_input_val - avg_buy_price) * abs(net_stand) if net_stand > 0 else (avg_sell_price - live_input_val) * abs(net_stand)
                pre_live_mtm_total += calc_mtm

    if metrics:
        total_cash_df = metrics["cash_df"] if metrics["cash_df"] is not None and not metrics["cash_df"].empty else pd.DataFrame()
        last_week_roll_val, pure_weekly_cash_flow, p_dep, p_wd = 0.0, 0.0, 0.0, 0.0
        
        if not total_cash_df.empty:
            roll_mask = total_cash_df['remarks'].astype(str).str.contains('Weekly Settlement Opening Entry', na=False, case=False)
            df_roll_entries = total_cash_df[roll_mask]
            close_mask = total_cash_df['remarks'].astype(str).str.contains('Weekly Settlement Closing Entry', na=False, case=False)
            df_pure_entries = total_cash_df[~roll_mask & ~close_mask]
            
            r_dep = float(df_roll_entries[df_roll_entries['tx_type'] == 'DEPOSIT']['amount'].sum())
            r_wd = float(df_roll_entries[df_roll_entries['tx_type'] == 'WITHDRAWAL']['amount'].sum())
            last_week_roll_val = r_dep - r_wd
            
            p_dep = float(df_pure_entries[df_pure_entries['tx_type'] == 'DEPOSIT']['amount'].sum())
            p_wd = float(df_pure_entries[df_pure_entries['tx_type'] == 'WITHDRAWAL']['amount'].sum())
            pure_weekly_cash_flow = p_dep - p_wd
            
        active_working_opening_base = last_week_roll_val if last_week_roll_val != 0 else metrics["opening_balance"]
        
        mc1, mc2, mc3, mc4, mc5 = st.columns(5)
        with mc1: st.markdown(f"""<div class="metric-card pastel-blue">Opening Capital Base<br><h2>₹ {active_working_opening_base:,.2f}</h2></div>""", unsafe_allow_html=True)
        with mc2: st.markdown(f"""<div class="metric-card pastel-yellow">Total Brokerage Friction<br><h2>₹ {metrics["total_brokerage"]:,.2f}</h2></div>""", unsafe_allow_html=True)
        with mc3:
            color_pnl = "pastel-green" if metrics["net_realized_pnl"] >= 0 else "pastel-red"
            st.markdown(f"""<div class="metric-card {color_pnl}">Net Realized P&L<br><h2>₹ {metrics["net_realized_pnl"]:,.2f}</h2></div>""", unsafe_allow_html=True)
        with mc4:
            color_mtm = "pastel-green" if pre_live_mtm_total >= 0 else "pastel-red"
            st.markdown(f"""<div class="metric-card {color_mtm}">Live Floating MTM Box<br><h2>₹ {pre_live_mtm_total:,.2f}</h2></div>""", unsafe_allow_html=True)
        with mc5:
            true_closing_capital = float(active_working_opening_base) + float(pure_weekly_cash_flow) + float(metrics["net_realized_pnl"]) + float(pre_live_mtm_total)
            color_final = "pastel-green" if true_closing_capital >= 0 else "pastel-red"
            st.markdown(f"""<div class="metric-card {color_final}">Final Current Balance<br><h2>₹ {true_closing_capital:,.2f}</h2></div>""", unsafe_allow_html=True)
            
        st.info(f"💵 **Current Week Capital Adjustments:** Deposits: ₹ {p_dep:,.2f} | Withdrawals: ₹ {p_wd:,.2f} (Net Cash Flow: ₹ {pure_weekly_cash_flow:+,.2f})")
        st.write("---")
# ==========================================
# [PART_22_END]
# ==========================================
# ==========================================
# [PART_23a_START] - Upgraded Multi-Client Consolidated Net Positions PDF Document Matrix Initializer
# ==========================================
        st.markdown("### 📥 Download Executive Statements & Specialized Accounts PDF Reports")
        
        # 🔥 MULTI CLIENT MASTER NET POSITION REPORT CREATOR ENGINE (AGGREGATE FIXED)
        def generate_all_clients_net_positions_pdf_matrix():
            m_buffer = io.BytesIO()
            m_doc = SimpleDocTemplate(
                m_buffer, pagesize=letter, rightMargin=20, 
                leftMargin=20, topMargin=25, bottomMargin=25
            )
            m_story = []
            m_styles = getSampleStyleSheet()
            
            t_style = ParagraphStyle(
                'MTTitle', parent=m_styles['Heading1'], fontSize=15, 
                textColor=colors.HexColor("#1e3a8a"), spaceAfter=15
            )
            s_style = ParagraphStyle(
                'MClientHead', parent=m_styles['Heading2'], fontSize=11, 
                textColor=colors.HexColor("#0f766e"), spaceBefore=14, 
                spaceAfter=6, fontName="Helvetica-Bold"
            )
            h_style = ParagraphStyle(
                'MHStyle', parent=m_styles['Normal'], fontSize=8, 
                leading=10, fontName="Helvetica-Bold", 
                textColor=colors.HexColor("#1e3a8a"), alignment=1
            )
            c_style = ParagraphStyle(
                'MCStyle', parent=m_styles['Normal'], fontSize=8, 
                leading=10, alignment=1
            )
            c_bold = ParagraphStyle(
                'MCBold', parent=m_styles['Normal'], fontSize=8, 
                leading=10, fontName="Helvetica-Bold", alignment=1
            )
            
            m_story.append(Paragraph(
                "SALASAR WEALTH - CONSOLIDATED NET POSITIONS MASTER REPORT", 
                t_style
            ))
            
            ts_label = datetime.now().strftime('%d-%b-%Y %H:%M')
            m_story.append(Paragraph(
                f"<b>Operational Week Block:</b> "
                f"{st.session_state['active_block']} | "
                f"<b>Generation Log:</b> {ts_label}", 
                ParagraphStyle('MSub', parent=m_styles['Normal'], fontSize=9)
            ))
            m_story.append(Spacer(1, 10))
            
            # 🔥 STEP 1: LOOP THROUGH ALL REGISTERED CLIENTS INDIVIDUALLY
            for client_lookup_name in CLIENTS:
                m_story.append(Paragraph(
                    f"👤 CLIENT PROFILE NET INVENTORY: "
                    f"{client_lookup_name.upper()}", s_style
                ))
                
                headers_pool = [
                    Paragraph("Scrip Name", h_style), 
                    Paragraph("Expiry Schedule", h_style), 
                    Paragraph("Net Open Qty", h_style), 
                    Paragraph("Avg Cost (₹)", h_style), 
                    Paragraph("Live Price (₹)", h_style), 
                    Paragraph("Floating MTM (₹)", h_style)
                ]
                
                client_specific_rows = [headers_pool]
                client_metrics_data = get_client_ledger_data(
                    client_lookup_name, st.session_state['active_block']
                )
                
                if client_metrics_data and not client_metrics_data["trades_df"].empty:
                    df_mc_pos = client_metrics_data["trades_df"].copy()
                    df_mc_pos['script_name'] = df_mc_pos['script_name'].astype(str).str.lower().str.strip()
                    df_mc_pos['selected_expiry'] = df_mc_pos['selected_expiry'].astype(str).str.lower().str.strip()
                    
                    # 🔥 CRITICAL AGGREGATE RE-MAP: Pehle poore block ka loop chalakar real volume net balance nikalenge
                    for (name_asset, expiry_asset), group_g in df_mc_pos.groupby(['script_name', 'selected_expiry']):
                        b_vol = int(group_g['buy_qty'].sum())
                        s_vol = int(group_g['sell_qty'].sum())
                        net_unsettled_volume = b_vol - s_vol
                        
                        if net_unsettled_volume != 0:
                            avg_buy_g = (group_g['buy_qty'] * group_g['buy_price']).sum() / b_vol if b_vol > 0 else 0.0
                            avg_sell_g = (group_g['sell_qty'] * group_g['sell_price']).sum() / s_vol if s_vol > 0 else 0.0
                            
                            live_p_rate = get_global_price(f"{name_asset}_{expiry_asset}")
                            if live_p_rate <= 0: 
                                live_p_rate = get_global_price(name_asset)
                            if live_p_rate <= 0:
                                live_p_rate = avg_buy_g if net_unsettled_volume > 0 else avg_sell_g
                                
                            if net_unsettled_volume > 0:
                                calculated_mtm = (live_p_rate - avg_buy_g) * abs(net_unsettled_volume)
                                label_direction = "BUY"
                                cost_rate = avg_buy_g
                            else:
                                calculated_mtm = (avg_sell_g - live_p_rate) * abs(net_unsettled_volume)
                                label_direction = "SELL"
                                cost_rate = avg_sell_g
                                
                            client_specific_rows.append([
                                Paragraph(name_asset.upper(), c_style),
                                Paragraph(expiry_asset.upper() if expiry_asset else "-", c_style), 
                                Paragraph(f"{abs(net_unsettled_volume)} [{label_direction}]", c_style),
                                Paragraph(f"{cost_rate:,.2f}", c_style), 
                                Paragraph(f"{live_p_rate:,.2f}", c_style),
                                Paragraph(f"{calculated_mtm:+,.2f}", c_bold)
                            ])
# ==========================================
# [PART_23a_END]
# ==========================================
# ==========================================
# [PART_23b_START] - PDF Report Compile Streams & Multi Client Interface Buttons
# ==========================================
                # 🔥 STEP 2: RENDER SEPARATE VISUAL TABLES SEGREGATED PER USER
                if len(client_specific_rows) > 1:
                    t_master_pos = Table(client_specific_rows, colWidths=None)
                    t_master_pos.setStyle(TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#eaeaea")), 
                        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
                        ('PADDING', (0,0), (-1,-1), 4), 
                        ('ALIGN', (0,0), (-1,-1), 'CENTER'), 
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
                    ]))
                    m_story.append(t_master_pos)
                else:
                    m_story.append(Paragraph(
                        "<i>🟢 Position Sheet Clear - No Active Standing Exposure.</i>", 
                        ParagraphStyle('MNo', parent=m_styles['Normal'], fontSize=8, 
                                       textColor=colors.HexColor("#475569"))
                    ))
                m_story.append(Spacer(1, 8))
                
            m_doc.build(m_story)
            m_buffer.seek(0)
            return m_buffer

        pdf_col1, pdf_col2, pdf_col3 = st.columns(3)
        with pdf_col1:
            pdf_data = generate_pdf_report(active_client, st.session_state['active_block'], metrics, pre_live_mtm_total)
            st.download_button(
                label="📥 Download Current Week Ledger Statement", 
                data=pdf_data, 
                file_name=f"Weekly_Ledger_{active_client}_{st.session_state['active_block']}.pdf", 
                mime="application/pdf", use_container_width=True
            )
        with pdf_col2:
            historical_pdf_data = generate_complete_historical_ledger_pdf(active_client)
            st.download_button(
                label="📜 Download All-Time Historical Statement", 
                data=historical_pdf_data, 
                file_name=f"Master_Historical_Ledger_{active_client}.pdf", 
                mime="application/pdf", use_container_width=True
            )
        with pdf_col3:
            net_pos_pdf_data = generate_only_net_positions_pdf(active_client, st.session_state['active_block'], metrics['trades_df'])
            st.download_button(
                label="📈 Download Only Active Net Positions Report", 
                data=net_pos_pdf_data, 
                file_name=f"Net_Positions_{active_client}_{st.session_state['active_block']}.pdf", 
                mime="application/pdf", use_container_width=True
            )
            
        st.write("")
        all_clients_pos_pdf_bytes = generate_all_clients_net_positions_pdf_matrix()
        
        fn_label = f"Consolidated_Master_Net_Positions_{st.session_state['active_block']}.pdf"
        st.download_button(
            label="📊 Download Multi-Client Consolidated Net Positions Master Sheet (Single File)",
            data=all_clients_pos_pdf_bytes,
            file_name=fn_label,
            mime="application/pdf",
            use_container_width=True,
            type="primary",
            key="master_all_clients_net_pos_download_btn"
        )
# ==========================================
# [PART_23b_END]
# ==========================================
# ==========================================
# [PART_24_BLOCK_A_START] - Tuple Unpack Dropdown Fix
# ==========================================
        st.write("---")
        st.markdown("### 📜 Sub-Broker Master Pack Report Center")
        try:
            conn_sb_tags = sqlite3.connect(
                'salasar_wealth_v19_ultimate.db'
            )
            cursor_sb_tags = conn_sb_tags.cursor()
            try:
                cursor_sb_tags.execute(
                    "UPDATE sub_broker_mappings "
                    "SET sub_broker_tag = UPPER(sub_broker_tag) "
                    "WHERE sub_broker_tag IS NOT NULL"
                )
                conn_sb_tags.commit()
            except Exception:
                pass
            cursor_sb_tags.execute(
                "SELECT DISTINCT UPPER(sub_broker_tag) "
                "FROM sub_broker_mappings "
                "WHERE sub_broker_tag IS NOT NULL "
                "AND sub_broker_tag != 'None'"
            )
            # 🔥 CRITICAL TUPLE FIX: Extracted string token using r[0] to clean brackets and commas completely
            sub_brokers_available_list = sorted(list(set([
                str(r[0]).strip().upper() 
                for r in cursor_sb_tags.fetchall() if r and r[0]
            ])))
            conn_sb_tags.close()
        except Exception:
            sub_brokers_available_list = []
            
        if not sub_brokers_available_list:
            st.info("ℹ️ Abhi tak koi Sub-Broker nahi hai.")
        else:
            selected_bulk_sb_target = st.selectbox(
                "🎯 Select Sub-Broker to Pack All Client PDFs Combined", 
                options=sub_brokers_available_list, 
                key="bulk_sb_download_selector_box"
            )
            if st.button(
                f"📦 Compile Consolidated Master Pack for "
                f"[{selected_bulk_sb_target.upper()}]", 
                use_container_width=True, 
                key="generate_combined_sb_pdf_btn"
            ):
                with st.spinner("Compiling multi-client ledgers..."):
                    try:
                        conn_fetch_group = sqlite3.connect(
                            'salasar_wealth_v19_ultimate.db'
                        )
                        cursor_fetch_group = conn_fetch_group.cursor()
                        cursor_fetch_group.execute(
                            "SELECT client_name FROM sub_broker_mappings "
                            "WHERE UPPER(sub_broker_tag) = UPPER(?)", 
                            (selected_bulk_sb_target,)
                        )
                        # 🔥 STRICT TEXT EXTRACTION LAYER APPLIED TO CLIENT NAMES AS WELL
                        associated_group_clients = sorted(list(set([
                            str(r[0]).strip() 
                            for r in cursor_fetch_group.fetchall() if r and r[0]
                        ])))
                        conn_fetch_group.close()
                        
                        if not associated_group_clients:
                            st.error("❌ No clients mapped under this sub-broker.")
                        else:
                            combined_buffer = io.BytesIO()
                            doc_combined = SimpleDocTemplate(
                                combined_buffer, pagesize=letter, 
                                rightMargin=20, leftMargin=20, 
                                topMargin=25, bottomMargin=25
                            )
                            story_master_pack = []
                            styles_m = getSampleStyleSheet()
                            t_st = ParagraphStyle(
                                'MT1', parent=styles_m['Heading1'], 
                                fontSize=15, textColor=colors.HexColor("#0d47a1"), 
                                spaceAfter=10
                            )
                            s_st = ParagraphStyle(
                                'MS1', parent=styles_m['Heading2'], 
                                fontSize=11, textColor=colors.HexColor("#1e88e5"), 
                                spaceBefore=14, spaceAfter=6, 
                                fontName="Helvetica-Bold"
                            )
                            sub_lbl = ParagraphStyle(
                                'MS2', parent=styles_m['Heading3'], 
                                fontSize=8.5, textColor=colors.HexColor("#475569"), 
                                spaceBefore=6, spaceAfter=4, 
                                fontName="Helvetica-Bold"
                            )
                            cell_st = ParagraphStyle(
                                'MC1', parent=styles_m['Normal'], 
                                fontSize=7.5, leading=9, alignment=1
                            )
                            cell_bd = ParagraphStyle(
                                'MC2', parent=styles_m['Normal'], 
                                fontSize=7.5, leading=9, 
                                fontName="Helvetica-Bold", alignment=1
                            )
                            h_st_main = ParagraphStyle(
                                'MH1', parent=styles_m['Normal'], 
                                fontSize=7.5, leading=9, 
                                fontName="Helvetica-Bold", 
                                textColor=colors.HexColor("#1e3a8a"), alignment=1
                            )
                            story_master_pack.append(Paragraph(
                                f"SALASAR WEALTH - CONSOLIDATED BATCH PACK", t_st
                            ))
                            story_master_pack.append(Paragraph(
                                f"<b>Sub-Broker Tag:</b> "
                                f"{selected_bulk_sb_target.upper()} | "
                                f"Block: {st.session_state['active_block']}", 
                                ParagraphStyle('MB1', parent=styles_m['Normal'], fontSize=9)
                            ))
                            story_master_pack.append(Spacer(1, 15))
# ==========================================
# [PART_24_BLOCK_A_END]
# ==========================================
# [PART_24_BLOCK_B_START]
# ==========================================
                            for sub_c in associated_group_clients:
                                c_m = get_client_ledger_data(
                                    sub_c, 
                                    st.session_state['active_block']
                                )
                                s_live_mtm = 0.0
                                if not c_m["trades_df"].empty:
                                    df_s_c = c_m["trades_df"].copy()
                                    df_s_c['script_name'] = df_s_c['script_name'].astype(str).str.lower().str.strip()
                                    df_s_c['selected_expiry'] = df_s_c['selected_expiry'].astype(str).str.strip()
                                    for (n_s, e_s), g_s in df_s_c.groupby(['script_name', 'selected_expiry']):
                                        b_v = int(g_s['buy_qty'].sum())
                                        s_v = int(g_s['sell_qty'].sum())
                                        n_st = b_v - s_v
                                        if n_st != 0:
                                            g_rate_s = get_global_price(f"{n_s}_{e_s}")
                                            if g_rate_s <= 0: g_rate_s = get_global_price(n_s)
                                            l_rate_s = float(g_rate_s if g_rate_s > 0 else ((g_s['buy_qty'] * g_s['buy_price']).sum() / b_v if n_st > 0 else (g_s['sell_qty'] * g_s['sell_price']).sum() / s_v))
                                            s_live_mtm += (l_rate_s - ((g_s['buy_qty'] * g_s['buy_price']).sum() / b_v)) * abs(n_st) if n_st > 0 else (((g_s['sell_qty'] * g_s['sell_price']).sum() / s_v) - l_rate_s) * abs(n_st)
                                
                                story_master_pack.append(Paragraph(f"👤 CLIENT LEDGER: {sub_c.upper()}", s_st))
                                story_master_pack.append(Paragraph("1. Performance Summary", sub_lbl))
                                t_c_df = c_m["cash_df"] if c_m["cash_df"] is not None and not c_m["cash_df"].empty else pd.DataFrame()
                                l_w_roll, p_weekly_flow = 0.0, 0.0
                                if not t_c_df.empty:
                                    r_mask = t_c_df['remarks'].astype(str).str.contains('Weekly Settlement Opening Entry', na=False, case=False)
                                    c_mask = t_c_df['remarks'].astype(str).str.contains('Weekly Settlement Closing Entry', na=False, case=False)
                                    df_p_e = t_c_df[~r_mask & ~c_mask]
                                    df_r_e = t_c_df[r_mask]
                                    l_w_roll = float(df_r_e[df_r_e['tx_type'] == 'DEPOSIT']['amount'].sum()) - float(df_r_e[df_r_e['tx_type'] == 'WITHDRAWAL']['amount'].sum())
                                    p_weekly_flow = float(df_p_e[df_p_e['tx_type'] == 'DEPOSIT']['amount'].sum()) - float(df_p_e[df_p_e['tx_type'] == 'WITHDRAWAL']['amount'].sum())
                                active_op_base = l_w_roll if l_w_roll != 0.0 else c_m["opening_balance"]
                                raw_trades_df = c_m["trades_df"]
                                raw_gross_realized_pnl = 0.0
                                total_brokerage_cost_friction = 0.0
                                if not raw_trades_df.empty:
                                    total_brokerage_cost_friction = float(raw_trades_df['brokerage'].sum())
                                    raw_gross_realized_pnl = float(raw_trades_df['manual_pnl'].sum())
                                f_close_s = float(active_op_base) + float(p_weekly_flow) + float(raw_gross_realized_pnl) + float(s_live_mtm) - float(total_brokerage_cost_friction)
                                s_data = [
                                    [Paragraph("Opening Base Capital Balance", cell_st), f"{active_op_base:,.2f}"],
                                    [Paragraph("Net Realized Trading P&L (Gross)", cell_st), f"{raw_gross_realized_pnl:,.2f}"],
                                    [Paragraph("Total Brokerage Friction (-)", cell_st), f"{total_brokerage_cost_friction:,.2f}"],
                                    [Paragraph("Live Floating MTM Box Total", cell_st), f"{s_live_mtm:,.2f}"],
                                    [Paragraph("Net Weekly Cash Adjustment (+/-)", cell_st), f"{p_weekly_flow:+,.2f}"],
                                    [Paragraph("<b>Final Current Balance Net Asset Value</b>", cell_bd), f"{f_close_s:,.2f}"]
                                ]
                                t_box = Table(s_data, colWidths=None)
                                t_box.setStyle(TableStyle([
                                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")), ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#f8fafc")),
                                    ('ALIGN', (1,0), (1,-1), 'RIGHT'), ('PADDING', (0,0), (-1,-1), 4)
                                ]))
                                story_master_pack.append(t_box)
# ==========================================
# [PART_24_BLOCK_B_END]
# ==========================================
# ==========================================
# [PART_24_BLOCK_C_START]
# ==========================================
                                story_master_pack.append(Paragraph("2. Weekly Capital Movements (Cash Logs)", sub_lbl))
                                if not c_m['cash_df'].empty:
                                    c_headers = [Paragraph("ID", h_st_main), Paragraph("Type", h_st_main), Paragraph("Amount Volume (₹)", h_st_main), Paragraph("Narration Remarks", h_st_main), Paragraph("Date & Time Log", h_st_main)]
                                    c_rows = [c_headers]
                                    for _, r_c in c_m['cash_df'].iterrows():
                                        clean_ts_str = str(r_c['timestamp'])[:16] if 'timestamp' in r_c and r_c['timestamp'] else "-"
                                        c_rows.append([str(r_c['id']), str(r_c['tx_type']), f"{r_c['amount']:,.2f}", str(r_c['remarks']), clean_ts_str])
                                    t_c_book = Table(c_rows, colWidths=None)
                                    t_c_book.setStyle(TableStyle([
                                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f1f5f9")), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#eaeaea")),
                                        ('PADDING', (0,0), (-1,-1), 4), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
                                    ]))
                                    story_master_pack.append(t_c_book)
                                else:
                                    story_master_pack.append(Paragraph("*No capital deposits or withdrawal entries logged.*", cell_st))

                                story_master_pack.append(Paragraph("3. Executed Trade Records Ledger Book", sub_lbl))
                                if not c_m['trades_df'].empty:
                                    tr_headers = [Paragraph("Scrip", h_st_main), Paragraph("Type", h_st_main), Paragraph("B_Qty", h_st_main), Paragraph("B_Price", h_st_main), Paragraph("S_Qty", h_st_main), Paragraph("S_Price", h_st_main), Paragraph("Brokerage", h_st_main), Paragraph("P&L (Gross)", h_st_main)]
                                    tr_rows = [tr_headers]
                                    for _, r_t in c_m['trades_df'].iterrows():
                                        b_qty_val = int(r_t['buy_qty'])
                                        s_qty_val = int(r_t['sell_qty'])
                                        b_price_val = float(r_t['buy_price'])
                                        s_price_val = float(r_t['sell_price'])
                                        line_gross_pnl = (s_price_val - b_price_val) * min(b_qty_val, s_qty_val) if b_qty_val > 0 and s_qty_val > 0 else 0.0
                                        tr_rows.append([str(r_t['script_name']).upper(), str(r_t['action_type']), str(b_qty_val), f"{b_price_val:,.2f}", str(s_qty_val), f"{s_price_val:,.2f}", f"{float(r_t['brokerage']):,.2f}", f"{line_gross_pnl:+,.2f}"])
                                    t_tr_book = Table(tr_rows, colWidths=None)
                                    t_tr_book.setStyle(TableStyle([
                                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f1f5f9")), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#eaeaea")),
                                        ('PADDING', (0,0), (-1,-1), 3), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
                                    ]))
                                    story_master_pack.append(t_tr_book)
                                else:
                                    story_master_pack.append(Paragraph("*No executed trades recorded in this block.*", cell_st))
                                    
                                story_master_pack.append(Paragraph("4. Standing Net Positions Inventory Book (Active Exposure)", sub_lbl))
                                p_rows = [[Paragraph("Scrip Name", h_st_main), Paragraph("Expiry", h_st_main), Paragraph("Net Open Qty", h_st_main), Paragraph("Avg Cost (₹)", h_st_main), Paragraph("Live Price (₹)", h_st_main), Paragraph("Floating MTM (₹)", h_st_main)]]
                                if not c_m['trades_df'].empty:
                                    df_sb_pos = c_m['trades_df'].copy()
                                    df_sb_pos['script_name'] = df_sb_pos['script_name'].astype(str).str.lower().str.strip()
                                    df_sb_pos['selected_expiry'] = df_sb_pos['selected_expiry'].astype(str).str.lower().str.strip()
                                    
                                    sb_totals_map = {}
                                    for (n_grp, e_grp), g_grp in df_sb_pos.groupby(['script_name', 'selected_expiry']):
                                        sb_totals_map[f"{n_grp}_{e_grp}"] = int(g_grp['buy_qty'].sum()) - int(g_grp['sell_qty'].sum())
                                        
                                    for _, r_p in df_sb_pos.iterrows():
                                        n_p = str(r_p['script_name']).lower().strip()
                                        e_p = str(r_p['selected_expiry']).lower().strip()
                                        if sb_totals_map.get(f"{n_p}_{e_p}", 0) != 0:
                                            live_p = get_global_price(f"{n_p}_{e_p}")
                                            if live_p <= 0: live_p = get_global_price(n_p)
                                            b_qty, s_qty, b_prc, s_prc = int(r_p['buy_qty']), int(r_p['sell_qty']), float(r_p['buy_price']), float(r_p['sell_price'])
                                            if live_p <= 0: live_p = b_prc if b_qty > 0 else s_prc
                                            if b_qty > 0 and s_qty == 0:
                                                c_mtm = (live_p - b_prc) * abs(b_qty)
                                                dir_lbl, cost_rt, disp_q = "BUY", b_prc, b_qty
                                            elif s_qty > 0 and b_qty == 0:
                                                c_mtm = (s_prc - live_p) * abs(s_qty)
                                                dir_lbl, cost_rt, disp_q = "SELL", s_prc, s_qty
                                            else:
                                                c_mtm = (live_p - b_prc) * abs(b_qty - s_qty)
                                                dir_lbl, cost_rt, disp_q = "BUY" if (b_qty - s_qty) > 0 else "SELL", b_prc, abs(b_qty - s_qty)
                                            
                                            chk_exist_list = [row_item[0] for row_item in p_rows]
                                            if n_p.upper() not in chk_exist_list:
                                                p_rows.append([n_p.upper(), e_p.upper() if e_p else "-", f"{abs(sb_totals_map[f'{n_p}_{e_p}'])} [{dir_lbl}]", f"{cost_rt:,.2f}", f"{live_p:,.2f}", f"{c_mtm:+,.2f}"])
                                if len(p_rows) > 1:
                                    t_sb_pos = Table(p_rows, colWidths=None)
                                    t_sb_pos.setStyle(TableStyle([
                                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#fff8e1")), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#eaeaea")),
                                        ('PADDING', (0,0), (-1,-1), 4), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
                                    ]))
                                    story_master_pack.append(t_sb_pos)
                                else:
                                    story_master_pack.append(Paragraph("*No active open standing positions recorded.*", cell_st))
                                story_master_pack.append(Spacer(1, 25))
                            doc_combined.build(story_master_pack)
                            combined_buffer.seek(0)
                            st.download_button(label="📥 Download Master Pack PDF", data=combined_buffer, file_name=f"Master_Pack_{selected_bulk_sb_target}.pdf", mime="application/pdf", use_container_width=True)
                            st.success("✨ Packed & Compiled Successfully!")
                    except Exception as err:
                        st.error(f"Compilation error: {str(err)}")
        st.write("---")
# ==========================================
# [PART_24_BLOCK_C_END]
# ==========================================
# ==========================================
# [PART_26_START] - Advanced Multi-Criteria Bulk Trades Cleanup Tool Panel
# ==========================================
        with st.expander("🗑️ Advanced Bulk Trades Wipe Tool (Client / Week / Date Wise)"):
            st.markdown("#### Filter Profiles to Delete Trade Records")
            st.warning("⚠️ Warning: Yeh feature sirf chuney huye records ko target karke clear karega, baaki sab safe rahega.")
            
            w_del_c1, w_del_c2, w_del_c3 = st.columns(3)
            with w_del_c1:
                target_del_client = st.selectbox("Select Target Profile to Wipe", CLIENTS, index=CLIENTS.index(active_client) if active_client in CLIENTS else 0, key="wipe_t_client_box")
            with w_del_c2:
                target_del_mode = st.selectbox("Delete Mode Scope", ["Selected Week Only", "Particular Specific Date", "All-Time Complete History"], key="wipe_t_mode_box")
            with w_del_c3:
                if target_del_mode == "Selected Week Only":
                    target_del_week = st.selectbox("Select Target Week Block", week_options, index=week_options.index(st.session_state['active_block']) if st.session_state['active_block'] in week_options else 0, key="wipe_t_week_box")
                    target_del_date_str = ""
                elif target_del_mode == "Particular Specific Date":
                    target_del_week = ""
                    target_del_date_str = st.text_input("Enter Target Date String (Exactly as logged)", placeholder="e.g. 25-06-2026", key="wipe_t_date_box")
                else:
                    target_del_week = ""
                    target_del_date_str = ""

            st.write("")
            confirm_bulk_wipe_checkbox = st.checkbox(f"I confirm that I want to wipe trades for **{target_del_client}** matching the selected scope criteria.", key="bulk_wipe_safety_lock_chk")
            
            if st.button(f"🔥 Execute Bulk Delete for {target_del_client}", type="primary", use_container_width=True, key="bulk_wipe_action_btn"):
                if not confirm_bulk_wipe_checkbox:
                    st.warning("⚠️ Action blocked! Pehle safety check confirmation box ko tick karein.")
                else:
                    conn_wipe = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                    cursor_wipe = conn_wipe.cursor()
                    
                    if target_del_mode == "All-Time Complete History":
                        cursor_wipe.execute("DELETE FROM trades WHERE client_name = ?", (target_del_client,))
                        success_msg = f"✨ Perfect! **{target_del_client}** ke saare historical records table se wipe out kar diye gaye hain."
                    elif target_del_mode == "Selected Week Only":
                        cursor_wipe.execute("DELETE FROM trades WHERE client_name = ? AND week_block = ?", (target_del_client, target_del_week))
                        success_msg = f"✨ Perfect! **{target_del_client}** ke week block `[{target_del_week}]` ke saare trades successfully delete ho gaye hain."
                    elif target_del_mode == "Particular Specific Date":
                        if not target_del_date_str.strip():
                            st.error("❌ Date missing! Pehle target text input field me entry date mention karein.")
                            conn_wipe.close()
                            st.stop()
                        cursor_wipe.execute("DELETE FROM trades WHERE client_name = ? AND timestamp LIKE ?", (target_del_client, f"{target_del_date_str.strip()}%"))
                        success_msg = f"✨ Perfect! **{target_del_client}** ke date `[{target_del_date_str.strip()}]` se match hone wale trades delete ho gaye hain."
                    
                    conn_wipe.commit()
                    conn_wipe.close()
                    st.success(success_msg)
                    st.rerun()
# ==========================================
# [PART_26_END]
# ==========================================
# ==========================================
# [PART_27_START] - Advanced WhatsApp Link Generator Data Extraction Layer
# ==========================================
        # -------------------------------------------------------------------------
        # 🔥 MULTI-OPTION 100% BLOCKER-FREE DIRECT CLICK ENGINE (PART 1)
        # -------------------------------------------------------------------------
        st.write("---")
        st.markdown("### 💬 Instant WhatsApp Dispatch Cockpit (Free Web Channel)")
        
        # Pull the absolute stored phone number coordinates dynamically for THIS active client
        conn_p_fetch = sqlite3.connect('salasar_wealth_v19_ultimate.db')
        cursor_p_fetch = conn_p_fetch.cursor()
        cursor_p_fetch.execute("SELECT whatsapp_phone FROM client_settings WHERE client_name = ?", (active_client,))
        phone_row_data = cursor_p_fetch.fetchone()
        conn_p_fetch.close()
        
        # CORE DIGITS PURGE ENGINE: Ensures absolutely clean phone strings to prevent Google search loops completely
        raw_phone_extracted = str(phone_row_data if phone_row_data and phone_row_data is not None else "").strip()
        clean_phone_str = ""
        for char in raw_phone_extracted:
            if char.isdigit():
                clean_phone_str += char
        
        w_col1, w_col2 = st.columns([0.4, 0.6])
        with w_col1:
            cust_phone = st.text_input("🟢 Client WhatsApp Phone Number", value=clean_phone_str, placeholder="e.g. 919876543210 (With Country Code)", key=f"whatsapp_phone_field_{active_client}")
        
        # Final absolute clean variable confirmation tracking
        final_clean_phone = "".join([c for c in str(cust_phone).strip() if c.isdigit()])
# ==========================================
# [PART_27_END]
# ==========================================
# ==========================================
# [PART_28_START] - WhatsApp Dynamic Text Format Splitter Matrix
# ==========================================
        with w_col2:
            positions_text_block_opt1 = ""
            positions_text_block_opt2 = ""
            positions_preview_html_opt1 = ""
            positions_preview_html_opt2 = ""
            
            if not metrics["trades_df"].empty:
                for (name, expiry), group in metrics["trades_df"].groupby(['script_name', 'selected_expiry']):
                    b_vol = int(group['buy_qty'].sum())
                    s_vol = int(group['sell_qty'].sum())
                    net_qty = b_vol - s_vol
                    
                    if net_qty != 0:
                        # Fetch prices normally for Option 2 usage
                        avg_b_p = (group['buy_qty'] * group['buy_price']).sum() / b_vol if b_vol > 0 else 0.0
                        avg_s_p = (group['sell_qty'] * group['sell_price']).sum() / s_vol if s_vol > 0 else 0.0
                        
                        g_rate = get_global_price(f"{name}_{expiry}")
                        if g_rate <= 0: g_rate = get_global_price(name)
                        if g_rate <= 0: g_rate = avg_b_p if net_qty > 0 else avg_s_p
                        
                        direction_badge = "BUY" if net_qty > 0 else "SELL"
                        emoji_badge = "🔵" if net_qty > 0 else "🔴"
                        exp_lbl = f" ({expiry})" if str(expiry).strip() else ""
                        
                        # 🔥 OPTION 1 LOGIC: Price completely removed from text and screen preview
                        positions_text_block_opt1 += f"{emoji_badge} {name.upper()}{exp_lbl}: {abs(net_qty)} Qty [{direction_badge}]\n"
                        
                        # 🔥 COLOR FIX FOR OPTION 1 PREVIEW LOOP: Green for BUY, Red for SELL
                        if net_qty > 0:
                            positions_preview_html_opt1 += f"<div>🔹 <span style='color:#28a745; font-weight:bold;'>{name.upper()}{exp_lbl}: {abs(net_qty)} Qty [BUY]</span></div>"
                        else:
                            positions_preview_html_opt1 += f"<div>🔺 <span style='color:#dc3545; font-weight:bold;'>{name.upper()}{exp_lbl}: {abs(net_qty)} Qty [SELL]</span></div>"
                        
                        # 🔥 OPTION 2 LOGIC: 100% Restored to old standard style with live price parameters
                        positions_text_block_opt2 += f"  {emoji_badge} {name.upper()}{exp_lbl}: {abs(net_qty)} Qty [{direction_badge}] @ {g_rate:,.2f}\n"
                        if net_qty > 0:
                            positions_preview_html_opt2 += f"<div>🔹 <span style='color:#28a745; font-weight:bold;'>{name.upper()}{exp_lbl}: {abs(net_qty)} Qty [BUY] @ {g_rate:,.2f}</span></div>"
                        else:
                            positions_preview_html_opt2 += f"<div>🔺 <span style='color:#dc3545; font-weight:bold;'>{name.upper()}{exp_lbl}: {abs(net_qty)} Qty [SELL] @ {g_rate:,.2f}</span></div>"
                        
            if not positions_text_block_opt1:
                positions_text_block_opt1 = "  🔹 No Active Open Standing Positions\n"
                positions_preview_html_opt1 = "<div>🔹 No Active Open Standing Positions</div>"
            if not positions_text_block_opt2:
                positions_text_block_opt2 = "  🔹 No Active Open Standing Positions\n"
                positions_preview_html_opt2 = "<div>🔹 No Active Open Standing Positions</div>"

            import urllib.parse
            # 🔥 OPTION 1 TEXT TEMPLATE: Clean view with absolutely NO price and NO Net Bal lines anywhere
            option_1_text = (
                f"Account Statement\n\n"
                f"Client: {active_client.upper()}\n"
                f"Week Block: {st.session_state['active_block']}\n\n"
                f"----------------------------\n"
                f"ACTIVE OPEN STANDING INVENTORY:\n"
                f"{positions_text_block_opt1}"
                f"----------------------------\n"
                f"Live system dynamic position log update."
            )
            url_args_option_1 = urllib.parse.quote(option_1_text)
            whatsapp_url_option_1 = f"whatsapp://send?phone={final_clean_phone}&text={url_args_option_1}"
            
            # 🔥 OPTION 2 TEXT TEMPLATE: 100% Original full style layout with Live Prices and Net Balance lines restored
            option_2_text = (
                f"Account Statement\n\n"
                f"Client Profile: {active_client.upper()}\n"
                f"Accounting Week: {st.session_state['active_block']}\n\n"
                f"----------------------------\n"
                f"Opening Capital Base: Rs {active_working_opening_base:,.2f}\n"
                f"Net Weekly Realized P&L: Rs {metrics['net_realized_pnl']:,.2f}\n"
                f"Active Open Floating MTM: Rs {pre_live_mtm_total:,.2f}\n"
                f"Net Cash Deposits/WD: Rs {pure_weekly_cash_flow:+,.2f}\n"
                f"----------------------------\n"
                f"STANDING NET POSITIONS INVENTORY:\n"
                f"{positions_text_block_opt2}"
                f"----------------------------\n"
                f"FINAL NET CURRENT BALANCE: Rs {true_closing_capital:,.2f}\n\n"
                f"System generated live cockpit summary log lines."
            )
            url_args_option_2 = urllib.parse.quote(option_2_text)
            whatsapp_url_option_2 = f"whatsapp://send?phone={final_clean_phone}&text={url_args_option_2}"
# ==========================================
# [PART_28_END]
# ==========================================
# ==========================================
# [PART_29_START] - WhatsApp Option Tabs Screen Previews Rendering
# ==========================================
            if final_clean_phone:
                st.markdown("##### 📲 Select Option to Send via Direct Click:")
                opt1_tab, opt2_tab = st.tabs(["📈 Option 1 (Net Positions)", "📋 Option 2 (Full Detailed Statement)"])
                
                with opt1_tab:
                    st.markdown("*Step 1: Visual Color-Coded Screen Preview:*")
                    with st.container(border=True):
                        st.markdown(f"**Client:** {active_client.upper()} | **Block:** {st.session_state['active_block']}")
                        st.markdown("**ACTIVE OPEN INVENTORY:**")
                        # 🔥 Renders pure Qty lines (Now with proper Green/Red color matrix) and hides Net Bal lines
                        st.markdown(positions_preview_html_opt1, unsafe_allow_html=True)
                    st.write("")
                    st.markdown(f'<div style="text-align:center;"><a href="{whatsapp_url_option_1}" style="text-decoration:none;"><div style="background-color:#1e3a8a; color:white; padding:12px; border-radius:8px; font-weight:bold; font-size:14px;">🚀 Option 1: Send Active Net Positions to WhatsApp App</div></a></div>', unsafe_allow_html=True)
                    
                with opt2_tab:
                    st.markdown("*Step 1: Visual Color-Coded Full Statement Preview:*")
                    with st.container(border=True):
                        st.markdown(f"**Client Profile:** {active_client.upper()} | Accounting Week: {st.session_state['active_block']}")
                        st.markdown(f"Opening Base: **Rs {active_working_opening_base:,.2f}** | Realized P&L: **Rs {metrics['net_realized_pnl']:,.2f}**")
                        st.markdown(f"Active MTM: **Rs {pre_live_mtm_total:,.2f}** | Net Cash Flow: **Rs {pure_weekly_cash_flow:+,.2f}**")
                        st.markdown("---")
                        st.markdown("**STANDING NET POSITIONS INVENTORY:**")
                        # 🔥 Renders old live price lines to keep original layout fully intact
                        st.markdown(positions_preview_html_opt2, unsafe_allow_html=True)
                        st.markdown("---")
                        st.markdown(f"<span style='font-size:15px; font-weight:bold;'>FINAL NET CURRENT BALANCE: Rs {true_closing_capital:,.2f}</span>", unsafe_allow_html=True)
                    st.write("")
                    st.markdown(f'<div style="text-align:center;"><a href="{whatsapp_url_option_2}" style="text-decoration:none;"><div style="background-color:#4b5563; color:white; padding:12px; border-radius:8px; font-weight:bold; font-size:14px;">📋 Option 2: Send Full Statement to WhatsApp App</div></a></div>', unsafe_allow_html=True)
            else:
                st.warning("⚠️ WhatsApp Cockpit locked! Please enter phone number in configuration panel.")
# ==========================================
# [PART_29_END]
# ==========================================
# ==========================================
# [PART_30_START] - Executed Trade Table View Rendering
# ==========================================
        st.write("---")
        st.markdown("### 📋 Executed Trade Ledger Book (Weekly Records)")
        if not metrics["trades_df"].empty:
            view_df = metrics["trades_df"].copy()
            view_df['script_name'] = view_df['script_name'].astype(str).str.upper().str.strip()
            view_df['selected_expiry'] = view_df['selected_expiry'].astype(str).str.strip()
            
            display_ledger_view = view_df[['id', 'script_name', 'selected_expiry', 'action_type', 'exchange', 'buy_qty', 'buy_price', 'sell_qty', 'sell_price', 'turnover', 'brokerage', 'manual_pnl', 'status', 'timestamp']].copy()
            display_ledger_view.columns = ['ID', 'Script Name', 'Expiry', 'Action Type', 'Exchange', 'Buy Qty', 'Buy Price', 'Sell Qty', 'Sell Price', 'Turnover', 'Brokerage', 'Realized P&L', 'Status', 'Timestamp']
            
            st.dataframe(
                display_ledger_view.style.format({
                    "Buy Price": "{:,.2f}", "Sell Price": "{:,.2f}",
                    "Turnover": "{:,.2f}", "Brokerage": "{:,.2f}", "Realized P&L": "{:,.2f}"
                }),
                use_container_width=True, hide_index=True
            )
            
            st.write("")
            st.markdown("#### ⚡ Quick Row Action Modification Center")
            available_row_ids_list = sorted(list(display_ledger_view['ID'].unique()))
            
            inline_col1, inline_col2 = st.columns([0.3, 0.7])
            with inline_col1:
                selected_inline_id = st.selectbox("🎯 Select Trade ID From Above Table to Fix", options=["-- SELECT ID TO MODIFY --"] + [int(i) for i in available_row_ids_list], key=f"quick_inline_id_selector_{active_client}")
# ==========================================
# [PART_30_END]
# ==========================================
# ==========================================
# [PART_31_START] - Inline Editor & Dynamic Zero Brokerage Toggle Checkbox
# ==========================================
            if selected_inline_id != "-- SELECT ID TO MODIFY --":
                conn_inline = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                df_inline_row = pd.read_sql_query("SELECT * FROM trades WHERE id = ?", conn_inline, params=(selected_inline_id,))
                conn_inline.close()
                
                if not df_inline_row.empty:
                    with st.container(border=True):
                        st.markdown(f"**📝 Modifying Record ID: `{selected_inline_id}`** | Asset: **{str(df_inline_row.loc[0, 'script_name']).upper()}**")
                        
                        inf_col1, inf_col2, inf_col3, inf_col4 = st.columns(4)
                        with inf_col1:
                            inline_bqty = st.number_input("Modify Buy Qty", value=int(df_inline_row.loc[0, 'buy_qty']), key=f"in_bq_{selected_inline_id}")
                            inline_bpr = st.number_input("Modify Buy Price", value=float(df_inline_row.loc[0, 'buy_price']), key=f"in_bp_{selected_inline_id}")
                        with inf_col2:
                            inline_sqty = st.number_input("Modify Sell Qty", value=int(df_inline_row.loc[0, 'sell_qty']), key=f"in_sq_{selected_inline_id}")
                            inline_spr = st.number_input("Modify Sell Price", value=float(df_inline_row.loc[0, 'sell_price']), key=f"in_sp_{selected_inline_id}")
                        with inf_col3:
                            current_row_expiry = str(df_inline_row.loc[0, 'selected_expiry']).strip()
                            if current_row_expiry not in expiry_options:
                                expiry_options.insert(0, current_row_expiry)
                            inline_expiry = st.selectbox("Modify Expiry", options=expiry_options, index=expiry_options.index(current_row_expiry), key=f"in_exp_{selected_inline_id}")
                            
                            # 🔥 NEW INJECTED LAYER: Zero Brokerage Checkbox right inside your red-marked blank zone
                            inline_current_brokerage = float(df_inline_row.loc[0, 'brokerage'])
                            is_already_zero = True if inline_current_brokerage == 0.0 else False
                            inline_zero_b_chk = st.checkbox("Apply Zero Brokerage for this Trade", value=is_already_zero, key=f"in_zero_b_{selected_inline_id}")
                            
                        with inf_col4:
                            inline_status = st.selectbox("Modify Status", ["CARRY FORWARD", "SQUARED OFF", "SETTLED & MATCHED"], index=0 if df_inline_row.loc[0, 'status'] == "CARRY FORWARD" else 1, key=f"in_st_{selected_inline_id}")
                            
                        btn_col1, btn_col2 = st.columns(2)
                        with btn_col1:
                            if st.button("🗃️ Save Inline Changes Now", use_container_width=True, type="secondary", key=f"in_save_btn_{selected_inline_id}"):
                                new_turnover = (inline_bqty * inline_bpr) + (inline_sqty * inline_spr)
                                new_pnl = (inline_sqty * inline_spr) - (inline_bqty * inline_bpr) if (inline_bqty > 0 and inline_sqty > 0) else 0.0
                                
                                # If checkbox is ticked, force brokerage to 0.0, otherwise recalculate based on client profile settings
                                if inline_zero_b_chk:
                                    new_brokerage = 0.0
                                else:
                                    conn_x = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                                    cursor_x = conn_x.cursor()
                                    cursor_x.execute("SELECT brokerage_type, brokerage_nse, brokerage_mcx, brokerage_gifty, per_lot_rate FROM client_settings WHERE client_name = ?", (active_client,))
                                    c_sett_x = cursor_x.fetchone()
                                    conn_x.close()
                                    
                                    b_type_x, b_nse_x, b_mcx_x, b_gifty_x, l_rate_x = c_sett_x if c_sett_x else ("Per Crore", 1000, 1000, 1000, 0.0)
                                    if b_type_x == "Per Lot":
                                        new_brokerage = (inline_bqty + inline_sqty) * l_rate_x
                                    else:
                                        calc_rate_x = b_nse_x if df_inline_row.loc[0, 'exchange'] == 'NSE' else (b_mcx_x if df_inline_row.loc[0, 'exchange'] == 'MCX' else b_gifty_x)
                                        new_brokerage = (new_turnover * calc_rate_x) / 10000000
                                    
                                current_laptop_time = get_laptop_time()
                                conn_up = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                                cursor_up = conn_up.cursor()
                                cursor_up.execute("""
                                    UPDATE trades SET buy_qty=?, buy_price=?, sell_qty=?, sell_price=?, selected_expiry=?, turnover=?, brokerage=?, manual_pnl=?, status=?, timestamp=? WHERE id=?
                                """, (inline_bqty, inline_bpr, inline_sqty, inline_spr, inline_expiry, new_turnover, new_brokerage, new_pnl, inline_status, current_laptop_time, selected_inline_id))
                                conn_up.commit()
                                conn_up.close()
                                st.success(f"✨ Record ID {selected_inline_id} successfully updated inline!")
                                st.rerun()
                        with btn_col2:
                            chk_inline_del = st.checkbox("Confirm Deletion Lock", key=f"in_del_chk_{selected_inline_id}")
                            if st.button("🗑️ Delete Record Inline", use_container_width=True, type="primary", key=f"in_del_btn_{selected_inline_id}"):
                                if chk_inline_del:
                                    conn_del = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                                    cursor_del = conn_del.cursor()
                                    cursor_del.execute("DELETE FROM trades WHERE id = ?", (selected_inline_id,))
                                    conn_del.commit()
                                    conn_del.close()
                                    st.error(f"💥 Record ID {selected_inline_id} deleted successfully!")
                                    st.rerun()
                                else:
                                    st.warning("⚠️ Pehle confirmation box tick karein!")
        else:
            st.info("Is active week block mein filhal is profile ke koi executed trading ledger record saved nahi hain.")
        st.write("---")
# ==========================================
# [PART_31_END]
# ==========================================
# ==========================================
# [PART_32_START] - Standing Net Positions Inventory Book & Partial Square-Off
# ==========================================
        st.write("---")
        st.markdown("### 📈 Standing Net Positions Inventory Book (Live Floating & Partial Square-Off Engine)")
        
        if not metrics["trades_df"].empty:
            pos_data = []
            df_pos_clean = metrics["trades_df"].copy()
            df_pos_clean['script_name'] = df_pos_clean['script_name'].astype(str).str.lower().str.strip()
            df_pos_clean['selected_expiry'] = df_pos_clean['selected_expiry'].astype(str).str.strip()
            
            for (name, expiry), group in df_pos_clean.groupby(['script_name', 'selected_expiry']):
                b_vol = int(group['buy_qty'].sum())
                s_vol = int(group['sell_qty'].sum())
                net_qty = b_vol - s_vol
                avg_b_price = (group['buy_qty'] * group['buy_price']).sum() / b_vol if b_vol > 0 else 0.0
                # 🔥 CRITICAL VARIABLE NAME FIX: Replaced s_v with s_vol to prevent system crashes
                avg_s_price = (group['sell_qty'] * group['sell_price']).sum() / s_vol if s_vol > 0 else 0.0
                
                if net_qty != 0:
                    saved_global_rate = get_global_price(f"{name}_{expiry}")
                    if saved_global_rate <= 0: saved_global_rate = get_global_price(name)
                    bhav_rate = float(saved_global_rate if saved_global_rate > 0 else (avg_b_price if net_qty > 0 else avg_s_price))
                    
                    pos_data.append({
                        "Script": name.upper(), "Expiry": expiry, "Net Qty": net_qty, 
                        "Avg Buy": avg_b_price, "Avg Sell": avg_s_price, "Price": bhav_rate
                    })
            
            if pos_data:
                pc1, pc2 = st.columns(2)
                with pc1:
                    st.markdown("##### ⚙️ Live Adjustments & Partial Square-Off Panel")
                    for p in pos_data:
                        sc_lower = p['Script'].lower()
                        exp_str_clean = p['Expiry']
                        unique_id_key = f"{sc_lower}_{exp_str_clean.replace('-', '_')}"
                        exp_lbl = f" ({exp_str_clean})" if exp_str_clean else ""
                        
                        st.markdown(f"**⚡ Symbol: {p['Script']}{exp_lbl}** (Net Open Qty: `{p['Net Qty']:+d}`)")
                        live_in = st.number_input(f"Global Price for {p['Script']}{exp_lbl}", value=p["Price"], key=f"lp_d_{unique_id_key}", step=0.05, label_visibility="collapsed")
                        
                        with st.expander(f"🔴 Square Off / Partial Exit - {p['Script']}{exp_lbl}"):
                            with st.form(f"sq_form_{unique_id_key}"):
                                scol1, scol2 = st.columns(2)
                                with scol1:
                                    sq_qty = st.number_input("Exit Quantity", min_value=1, max_value=abs(p['Net Qty']), step=1, key=f"sq_q_{unique_id_key}")
                                with scol2:
                                    sq_price = st.number_input("Exit Price (₹)", min_value=0.0, value=live_in, step=0.05, key=f"sq_p_{unique_id_key}")
                                
                                if st.form_submit_button("🎯 Execute Square Off", use_container_width=True, type="primary"):
                                    conn_b = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                                    cursor_b = conn_b.cursor()
                                    cursor_b.execute("SELECT brokerage_type, brokerage_nse, brokerage_mcx, brokerage_gifty, per_lot_rate FROM client_settings WHERE client_name = ?", (active_client,))
                                    c_sett = cursor_b.fetchone()
                                    conn_b.close()
                                    
                                    b_type, b_nse, b_mcx, b_gifty, l_rate = c_sett if c_sett else ("Per Crore", 1000, 1000, 1000, 0.0)
                                    calc_turnover = sq_qty * sq_price
                                    calc_brokerage = sq_qty * l_rate if b_type == "Per Lot" else (calc_turnover * b_nse) / 10000000
                                        
                                    current_laptop_time = get_laptop_time()
                                    conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                                    cursor = conn.cursor()
                                    
                                    if p['Net Qty'] > 0:
                                        cursor.execute("""
                                            INSERT INTO trades (client_name, week_block, exchange, script_name, selected_expiry, action_type, buy_qty, buy_price, sell_qty, sell_price, turnover, brokerage, manual_pnl, status, timestamp)
                                            VALUES (?, ?, 'NSE', ?, ?, 'SELL', 0, 0.0, ?, ?, ?, ?, 0.0, 'SQUARED OFF', ?)
                                        """, (active_client, st.session_state['active_block'], sc_lower, exp_str_clean, sq_qty, sq_price, calc_turnover, calc_brokerage, current_laptop_time))
                                    else:
                                        cursor.execute("""
                                            INSERT INTO trades (client_name, week_block, exchange, script_name, selected_expiry, action_type, buy_qty, buy_price, sell_qty, sell_price, turnover, brokerage, manual_pnl, status, timestamp)
                                            VALUES (?, ?, 'NSE', ?, ?, 'BUY', ?, ?, 0, 0.0, ?, ?, 0.0, 'SQUARED OFF', ?)
                                        """, (active_client, st.session_state['active_block'], sc_lower, exp_str_clean, sq_qty, sq_price, calc_turnover, calc_brokerage, current_laptop_time))
                                    conn.commit()
                                    conn.close()
                                    st.success(f"✅ Squared off {sq_qty} qty for {p['Script']}!")
                                    st.rerun()
                        st.write("") 
                        
                with pc2:
                    st.markdown("##### 📊 Floating MTM Breakup Overview")
                    for p in pos_data:
                        unique_id_key = f"{p['Script'].lower()}_{p['Expiry'].replace('-', '_')}"
                        live_in = st.session_state.get(f"lp_d_{unique_id_key}", p["Price"])
                        sc_mtm = (live_in - p["Avg Buy"]) * abs(p["Net Qty"]) if p["Net Qty"] > 0 else (p["Avg Sell"] - live_in) * abs(p["Net Qty"])
                        
                        icon = "🟢" if sc_mtm >= 0 else "🔴"
                        exp_lbl = f" ({p['Expiry']})" if p['Expiry'] else ""
                        st.markdown(f"{icon} **{p['Script']}{exp_lbl} MTM:** ₹ {sc_mtm:,.2f} &nbsp;&nbsp;&nbsp;&nbsp; `[ Qty: {p['Net Qty']:+d} ]`")
            else:
                st.info("No active open exposure records carried forward.")
        else:
            st.info("Inventory balance sheet clear.")
# ==========================================
# [PART_32_END]
# ==========================================
# ==========================================
# [PART_33] - Global Price Controller & Bulk System Engine (Complete)
# ==========================================
with tab4:
    st.subheader("🌐 Global Price Sync & System-Wide Bulk Settlement Center")
    st.info("Yahan se aap unique scrip + expiry ke universal rates set karke automatic bulk settlement run kar sakte hain.")
    
    conn_global = sqlite3.connect('salasar_wealth_v19_ultimate.db')
    # 🔥 CRITICAL FIX: Filtering records STRICTLY by active week block instead of scanning all-time history tables
    df_all_trades_raw = pd.read_sql_query(
        "SELECT script_name, selected_expiry, buy_qty, sell_qty "
        "FROM trades WHERE week_block = ?", 
        conn_global, 
        params=(st.session_state['active_block'],)
    )
    conn_global.close()
    
    active_global_symbols = set()
    if not df_all_trades_raw.empty:
        df_all_trades_raw['script_name'] = df_all_trades_raw['script_name'].astype(str).str.lower().str.strip()
        # 🔥 FORCE EXPIRY UNIQUE KEY PARSING TO PREVENT DUPLICATE GENERATION COLOUMNS
        df_all_trades_raw['selected_expiry'] = df_all_trades_raw['selected_expiry'].astype(str).str.lower().str.strip()
        
        for (name, expiry), group in df_all_trades_raw.groupby(['script_name', 'selected_expiry']):
            if (int(group['buy_qty'].sum()) - int(group['sell_qty'].sum())) != 0:
                exp_suffix = f"_{expiry}" if expiry else ""
                active_global_symbols.add(f"{name}{exp_suffix}")
                
    if not active_global_symbols:
        active_global_symbols = ["nifty_30-jun-2026", "delhivery_30-jun-2026", "goldm", "sm_30-jun-2026", "sm_30-jul-2026"]
        
    # --- STEP 1: UNIVERSAL RATES SYNC PANEL ---
    with st.form("global_system_price_form"):
        st.markdown("#### 📝 Step 1: Set Universal Expiry-Wise Closing Rates")
        updated_prices = {}
        for symbol_key in sorted(active_global_symbols):
            current_rate = get_global_price(symbol_key)
            display_label = symbol_key.upper().replace("_", " (") + (")" if "_" in symbol_key else "")
            
            updated_prices[symbol_key] = st.number_input(
                f"Closing Price for [{display_label}] (₹)", 
                value=float(current_rate if current_rate > 0 else 100.0),
                step=0.05, key=f"settle_rate_{symbol_key}"
            )
        
        submit_prices = st.form_submit_button("🔥 Save & Sync Rates Across All Portfolios", use_container_width=True)
        if submit_prices:
            for sym, prc in updated_prices.items():
                update_global_price(sym, prc)
            st.success("Universal expiry-wise rates locked successfully!")
            st.rerun()
            
    st.write("---")
    st.markdown("#### ⚡ Step 2: System-Wide Bulk Weekly Settlement Engine")
    next_weeks_avail_bulk = [w for w in week_options if week_options.index(w) > week_options.index(st.session_state['active_block'])]
    target_bulk_week = st.selectbox("🎯 Select Next Target Week For BULK Carry Forward", next_weeks_avail_bulk if next_weeks_avail_bulk else ["No Future Weeks Created"], key="bulk_settle_wk_drop")
    confirm_bulk_checkbox = st.checkbox("Confirm Executing Bulk Settlement for ALL Clients simultaneously", key="cbc_lock_final")
    
    if st.button("🚀 Run System-Wide Bulk Weekly Settlement", use_container_width=True, type="primary", key="master_bulk_settle_btn"):
        if target_bulk_week == "No Future Weeks Created":
            st.error("❌ Bulk Settlement failed. Next week block sidebar me banayein!")
        elif not confirm_bulk_checkbox:
            st.warning("⚠️ Action blocked! Check bulk confirmation box first.")
        else:
            with st.spinner("Processing system-wide transactions..."):
                for client_profile_name in CLIENTS:
                    execute_advanced_weekly_settlement(client_profile_name, st.session_state['active_block'], target_bulk_week)
            st.balloons(); st.success(f"🎯 Master Bulk Settlement Executed perfectly inside '{target_bulk_week}'!"); st.rerun()
            
    st.write("---")
    st.markdown("#### ⚡ Step 3: Undo / Rollback Last Wrong Weekly Settlement")
    st.warning("⚠️ Warning: Yeh action naye week block se automatic carry forward trades ko completely wipe out kar dega.")
    rollback_target_week = st.selectbox("Select the Wrong Target Week to Rollback", week_options, key="rb_wk_drop_v20")
    confirm_rollback_check = st.checkbox("Confirm Permanent Settlement Rollback & Balance Reversal Lock", key="c_rb_chk_v20")
    
    if st.button("⏪ Execute System-Wide Settlement Rollback & Undo Data", use_container_width=True, type="secondary", key="master_undo_settle_btn"):
        if not confirm_rollback_check:
            st.warning("⚠️ Action blocked! Check rollback safety confirmation box first.")
        else:
            with st.spinner("Undoing settlement logs..."):
                conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                cursor = conn.cursor()
                cursor.execute("DELETE FROM trades WHERE week_block = ?", (rollback_target_week,))
                cursor.execute("DELETE FROM cash_transactions WHERE remarks LIKE '%Weekly Settlement%'")
                conn.commit(); conn.close()
                st.success("🔄 Master Settlement Rollback Executed Flawlessly!"); st.rerun()

    st.write("---")
    st.markdown("#### 📂 Step 4: Secure Database Snapshot One-Click Restore Vault")
    mc_b1, mc_b2 = st.columns([0.4, 0.6])
    with mc_b1:
        if st.button("💾 Create Instant Force Backup Now", use_container_width=True, type="secondary", key="force_instant_backup_btn_test"):
            execute_database_daily_backup(); st.success("✨ Today's database copy saved into vault!"); st.rerun()
            
    backup_vault_dir = "salasar_db_vault"
    available_restore_files = sorted([f for f in os.listdir(backup_vault_dir) if f.endswith('.db')], reverse=True) if os.path.exists(backup_vault_dir) else []
    
    if not available_restore_files:
        st.warning("⚠️ Storage backup vault empty hai.")
    else:
        st.write("")
        selected_snapshot_to_restore = st.selectbox("⏳ Select Historical Target Snapshot To Restore:", available_restore_files, key="db_restore_selector_dropdown")
        st.warning(f"❗ Warning: Restore execute karte hi choice file '{selected_snapshot_to_restore}' live production par override ho jayegi.")
        confirm_dangerous_restore_chk = st.checkbox("Confirm Live Production Database Override Lock", key="chk_security_lock_restore_dangerous")
        
        if st.button("🔄 Execute One-Click Database Snapshot Restore", use_container_width=True, type="primary", key="force_restore_snapshot_btn"):
            if not confirm_dangerous_restore_chk: st.warning("⚠️ Operational action blocked! Pehle checkbox confirm kijiye.")
            else:
                with st.spinner("Restoring snapshots..."):
                    try:
                        import shutil
                        shutil.copy2(os.path.join(backup_vault_dir, selected_snapshot_to_restore), 'salasar_wealth_v19_ultimate.db')
                        st.balloons(); st.success("✨ Database Snapshot Restored perfectly!"); st.rerun()
                    except Exception as e: st.error(f"Error: {str(e)}")
# ==========================================
# [PART_33_END]
# ==========================================
# ==========================================
# [PART_34_START] - Upgraded Expiry Pivot View & Symmetrical Combined Dashboard
# ==========================================
with tab5:
    st.subheader("📊 Combined Multi-Client Snapshot Dashboard")
    st.info("System ke sabhi accounts ka data aur unique asset + expiry standing volumes live tracked hain.")
    
    snapshot_rows = []
    matrix_scrip_data = {}
    
    for c_name in CLIENTS:
        c_metrics = get_client_ledger_data(c_name, st.session_state['active_block'])
        c_live_mtm_total = 0.0
        
        if c_metrics and not c_metrics["trades_df"].empty:
            df_snapshot_trades = c_metrics["trades_df"][c_metrics["trades_df"]['week_block'] == st.session_state['active_block']].copy()
            df_snapshot_trades['script_name'] = df_snapshot_trades['script_name'].astype(str).str.lower().str.strip()
            # 🔥 FORCE CASE INSENSITIVE ALIGNMENT INSIDE ENTERPRISE DASHBOARD LOOP
            df_snapshot_trades['selected_expiry'] = df_snapshot_trades['selected_expiry'].astype(str).str.lower().str.strip()
            
            for (s_name, s_expiry), group_g in df_snapshot_trades.groupby(['script_name', 'selected_expiry']):
                b_v = int(group_g['buy_qty'].sum())
                s_v = int(group_g['sell_qty'].sum())
                n_stand_g = b_v - s_v
                
                if n_stand_g != 0:
                    exp_lbl = f" ({s_expiry.upper()})" if s_expiry else " (BASE)"
                    s_upper_compound = f"{s_name.upper().strip()}{exp_lbl}"
                    
                    if s_upper_compound not in matrix_scrip_data:
                        matrix_scrip_data[s_upper_compound] = {}
                    matrix_scrip_data[s_upper_compound][c_name] = n_stand_g
                    
                    avg_buy_g = (group_g['buy_qty'] * group_g['buy_price']).sum() / b_v if b_v > 0 else 0.0
                    avg_sell_g = (group_g['sell_qty'] * group_g['sell_price']).sum() / s_v if s_v > 0 else 0.0
                    
                    global_saved_rate = get_global_price(f"{s_name}_{s_expiry}")
                    if global_saved_rate <= 0:
                        global_saved_rate = get_global_price(s_name)
                    live_input_val = float(global_saved_rate if global_saved_rate > 0 else (avg_buy_g if n_stand_g > 0 else avg_sell_g))
                    
                    calc_mtm = (live_input_val - avg_buy_g) * abs(n_stand_g) if n_stand_g > 0 else (avg_sell_g - live_input_val) * abs(n_stand_g)
                    c_live_mtm_total += calc_mtm
                    
        total_cash_df = c_metrics["cash_df"] if c_metrics and not c_metrics["cash_df"].empty else pd.DataFrame()
        last_week_roll_val = 0.0
        pure_weekly_cash_flow = 0.0
        
        if not total_cash_df.empty:
            roll_mask = total_cash_df['remarks'].astype(str).str.contains('Weekly Settlement Opening Entry', na=False, case=False)
            df_roll_entries = total_cash_df[roll_mask]
            close_mask = total_cash_df['remarks'].astype(str).str.contains('Weekly Settlement Closing Entry', na=False, case=False)
            df_pure_entries = total_cash_df[~roll_mask & ~close_mask]
            
            r_dep = float(df_roll_entries[df_roll_entries['tx_type'] == 'DEPOSIT']['amount'].sum())
            r_wd = float(df_roll_entries[df_roll_entries['tx_type'] == 'WITHDRAWAL']['amount'].sum())
            last_week_roll_val = r_dep - r_wd
            
            p_dep = float(df_pure_entries[df_pure_entries['tx_type'] == 'DEPOSIT']['amount'].sum())
            p_wd = float(df_pure_entries[df_pure_entries['tx_type'] == 'WITHDRAWAL']['amount'].sum())
            pure_weekly_cash_flow = p_dep - p_wd
            
        c_realized = c_metrics["net_realized_pnl"] if c_metrics else 0.0
        
        conn_snapshot_fix = sqlite3.connect('salasar_wealth_v19_ultimate.db')
        cursor_snapshot_fix = conn_snapshot_fix.cursor()
        cursor_snapshot_fix.execute("SELECT opening_balance FROM client_settings WHERE client_name = ?", (c_name,))
        row_snapshot_fix = cursor_snapshot_fix.fetchone()
        conn_snapshot_fix.close()
        
        c_opening = float(row_snapshot_fix[0]) if row_snapshot_fix and row_snapshot_fix is not None else 0.0
        active_display_opening_base = last_week_roll_val if last_week_roll_val != 0 else c_opening
        c_final_current_bal = float(active_display_opening_base) + float(pure_weekly_cash_flow) + float(c_realized) + float(c_live_mtm_total)
        
        snapshot_rows.append({
            "Client Name": c_name, "Opening Capital Base (₹)": active_display_opening_base,
            "Last Week Rollover Balance (₹)": last_week_roll_val, "Net Cash Flow (₹)": pure_weekly_cash_flow,
            "Realized Trading P&L (₹)": c_realized, "Live Floating MTM (₹)": c_live_mtm_total,
            "Final Current Balance (₹)": c_final_current_bal
        })
        
    df_snapshot = pd.DataFrame(snapshot_rows)
    st.dataframe(df_snapshot.style.format({
        "Opening Capital Base (₹)": "{:,.2f}", "Last Week Rollover Balance (₹)": "{:,.2f}",
        "Net Cash Flow (₹)": "{:+,.2f}", "Realized Trading P&L (₹)": "{:,.2f}",
        "Live Floating MTM (₹)": "{:,.2f}", "Final Current Balance (₹)": "{:,.2f}"
    }), use_container_width=True, hide_index=True)
    
    st.markdown("#### 📦 Aggregate Standing Volume Matrix (All Portfolios Combined)")
    if matrix_scrip_data:
        matrix_rows = []
        for scrip_key, client_vols in matrix_scrip_data.items():
            row_dict = {"Script & Expiry Name": scrip_key}
            total_scrip_vol = 0
            for client_col in CLIENTS:
                qty_val = client_vols.get(client_col, 0)
                row_dict[client_col] = qty_val
                total_scrip_vol += qty_val
            row_dict["Total System Volume"] = total_scrip_vol
            matrix_rows.append(row_dict)
            
        df_pivot_matrix = pd.DataFrame(matrix_rows)
        fmt_config_matrix = {}
        for client_col in CLIENTS:
            fmt_config_matrix[client_col] = lambda x: f"{x:+d}" if x != 0 else "0"
        fmt_config_matrix["Total System Volume"] = lambda x: f"{x:+d}" if x != 0 else "0"
        
        st.dataframe(df_pivot_matrix.style.format(fmt_config_matrix), use_container_width=True, hide_index=True)
    else:
        st.info("Filhal pooré system mein kisi bhi client ki koi open standing position nahi hai.")
# ==========================================
# [PART_34_END]
# ==========================================
# ==========================================
# [PART_35_START] - Advanced Trade Editor & Log Fixer Matrix
# ==========================================
    st.write("---")
    with st.expander("🛠️ Advanced Entry Modification & Editing Panel (Fix & Delete Trades)"):
        mod_id = st.number_input("Enter exact Trade Database ID to Modify/Delete", min_value=1, step=1, key="mod_field_id_final")
        
        if mod_id:
            conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
            df_curr = pd.read_sql_query("SELECT * FROM trades WHERE id = ?", conn, params=(mod_id,))
            conn.close()
            
            if not df_curr.empty:
                with st.form(f"mod_form_{mod_id}_final"):
                    st.markdown(f"#### 📝 Modifying Trade Log ID: `{mod_id}`")
                    m_bqty = st.number_input("New Buy Qty", value=int(df_curr.loc[0, 'buy_qty']))
                    m_bpr = st.number_input("New Buy Price", value=float(df_curr.loc[0, 'buy_price']))
                    m_sqty = st.number_input("New Sell Qty", value=int(df_curr.loc[0, 'sell_qty']))
                    m_spr = st.number_input("New Sell Price", value=float(df_curr.loc[0, 'sell_price']))
                    
                    # 🔥 NEW ENTRY FIELD: Dynamic Expiry Change Selectbox Dropdown Matrix
                    current_saved_expiry = str(df_curr.loc[0, 'selected_expiry']).strip()
                    # Ensures pre-existing target item stays safe inside options matrix list arrays
                    if current_saved_expiry not in expiry_options:
                        expiry_options.insert(0, current_saved_expiry)
                    m_expiry = st.selectbox("Modify Selected Expiry", options=expiry_options, index=expiry_options.index(current_saved_expiry), key=f"me_{mod_id}")
                    
                    m_status = st.selectbox("New Status", ["CARRY FORWARD", "SQUARED OFF", "SETTLED & MATCHED"], index=0 if df_curr.loc[0, 'status'] == "CARRY FORWARD" else 1)
                    
                    b_update, b_delete = st.columns(2)
                    with b_update:
                        if st.form_submit_button("🗃️ Force Update Record Attributes", use_container_width=True):
                            new_turnover = (m_bqty * m_bpr) + (m_sqty * m_spr)
                            new_pnl = (m_sqty * m_spr) - (m_bqty * m_bpr) if (m_bqty > 0 and m_sqty > 0) else 0.0
                            
                            conn_x = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                            cursor_x = conn_x.cursor()
                            cursor_x.execute("SELECT brokerage_type, brokerage_nse, brokerage_mcx, brokerage_gifty, per_lot_rate FROM client_settings WHERE client_name = ?", (active_client,))
                            c_sett_x = cursor_x.fetchone()
                            conn_x.close()
                            
                            b_type_x, b_nse_x, b_mcx_x, b_gifty_x, l_rate_x = c_sett_x if c_sett_x else ("Per Crore", 1000, 1000, 1000, 0.0)
                            if b_type_x == "Per Lot":
                                new_brokerage = (m_bqty + m_sqty) * l_rate_x
                            else:
                                calc_rate_x = b_nse_x if df_curr.loc[0, 'exchange'] == 'NSE' else (b_mcx_x if df_curr.loc[0, 'exchange'] == 'MCX' else b_gifty_x)
                                new_brokerage = (new_turnover * calc_rate_x) / 10000000
                                
                            current_laptop_time = get_laptop_time()
                            conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                            cursor = conn.cursor()
                            # 🔥 FIXED STATEMENT: Added selected_expiry column update handler target
                            cursor.execute("""
                                UPDATE trades SET buy_qty=?, buy_price=?, sell_qty=?, sell_price=?, selected_expiry=?, turnover=?, brokerage=?, manual_pnl=?, status=?, timestamp=? WHERE id=?
                            """, (m_bqty, m_bpr, m_sqty, m_spr, m_expiry, new_turnover, new_brokerage, new_pnl, m_status, current_laptop_time, mod_id))
                            conn.commit(); conn.close(); st.success("Entry modified successfully!"); st.rerun()
                    with b_delete:
                        chk_del_trade = st.checkbox("Confirm Permanent Trade Deletion (Safety Locked)", key=f"tdc_{mod_id}")
                        if st.form_submit_button("🗑️ Delete This Trade Permanently", use_container_width=True, type="primary"):
                            if chk_del_trade:
                                conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
                                cursor = conn.cursor()
                                cursor.execute("DELETE FROM trades WHERE id = ?", (mod_id,))
                                conn.commit(); conn.close(); st.error(f"💥 Trade ID {mod_id} Deleted!"); st.rerun()
                            else: st.warning("⚠️ Action blocked! Please check the safety confirmation box first.")
            else:
                st.warning("⚠️ Specified Trade Database ID not found inside database logs.")
# ==========================================
# [PART_35_END]
# ==========================================
# ==========================================
# [PART_36_START] - Standalone Tab 6 Excel Style Scrip Auditing Engine
# ==========================================
with tab6:
    st.subheader("🔍 Excel-Style Scrip Wise Audit & Net Profit Book")
    st.info("Yahan par aap kisi bhi single client ke specific scrip asset ko select karke uske automatic dynamic calculations aur custom row highlights track kar sakte hain.")
    
    # Base calculation pool pull from the active client ledger
    t6_metrics = get_client_ledger_data(active_client, st.session_state['active_block'])
    
    if t6_metrics and not t6_metrics["trades_df"].empty:
        all_t6_logs_df = t6_metrics["trades_df"].copy()
        all_t6_logs_df['script_name'] = all_t6_logs_df['script_name'].astype(str).str.lower().str.strip()
        all_t6_logs_df['selected_expiry'] = all_t6_logs_df['selected_expiry'].astype(str).str.lower().str.strip()
        
        # Extract dynamic unique symbols alphabetical sorting
        t6_scrips_list = sorted(list(all_t6_logs_df['script_name'].str.upper().unique()))
        
        sc1, sc2 = st.columns([0.4, 0.6])
        with sc1:
            selected_t6_scrip = st.selectbox(
                f"📊 Select Asset to Filter for [ {active_client} ]:",
                ["SHOW ALL SCRIPS"] + t6_scrips_list,
                key="tab6_standalone_scrip_filter_dropdown"
            )
            
        # Applying dynamic query slices
        if selected_t6_scrip != "SHOW ALL SCRIPS":
            filtered_t6_df = all_t6_logs_df[all_t6_logs_df['script_name'].str.upper() == selected_t6_scrip].copy()
        else:
            filtered_t6_df = all_t6_logs_df.copy()
            
        # 🔥 STEP 1: CUMULATIVE VOLUMES NET ASSESSMENT ENGINE (EXPIRY-WISE AGGREGATION LAYER)
        scrip_net_positions_tracker = {}
        for (name, expiry), group in filtered_t6_df.groupby(['script_name', 'selected_expiry']):
            total_b_vol = int(group['buy_qty'].sum())
            total_s_vol = int(group['sell_qty'].sum())
            net_inventory_balance = total_b_vol - total_s_vol
            
            compound_key = f"{name}_{expiry}"
            scrip_net_positions_tracker[compound_key] = {
                "net_qty": net_inventory_balance,
                "total_buy_qty": total_b_vol,
                "total_sell_qty": total_s_vol
            }

        # 🔥 STEP 2: INDIVIDUAL ROW-WISE MATHEMATICAL CALCULATION MATRIX
        live_prices_list = []
        floating_mtm_list = []
        adjusted_final_pnl_list = []
        row_visual_status_tag = []
        
        for idx, row in filtered_t6_df.iterrows():
            r_name = str(row['script_name']).lower().strip()
            r_expiry = str(row['selected_expiry']).lower().strip()
            
            b_qty = int(row['buy_qty'])
            b_prc = float(row['buy_price'])
            s_qty = int(row['sell_qty'])
            s_prc = float(row['sell_price'])
            r_pnl = float(row['manual_pnl'])
            r_brok = float(row['brokerage'])
            
            compound_key = f"{r_name}_{r_expiry}"
            position_metrics = scrip_net_positions_tracker.get(compound_key, {"net_qty": 0})
            net_unsettled_volume = position_metrics["net_qty"]
            
            # Fetch prices normally from universal price repository
            global_saved_rate = get_global_price(f"{r_name}_{r_expiry}")
            if global_saved_rate <= 0:
                global_saved_rate = get_global_price(r_name)
            live_rate_current = float(global_saved_rate if global_saved_rate > 0 else (b_prc if b_qty > 0 else s_prc))
            
            # 🔥 CONDITION FIX: If cumulative expiry sum is completely matched, row is pure settled
            if net_unsettled_volume == 0:
                live_prices_list.append("-")
                floating_mtm_list.append("0.00 [SETTLED]")
                adjusted_final_pnl_list.append(r_pnl - r_brok)
                row_visual_status_tag.append("CLOSED")
            else:
                # 🔥 ROW-WISE EXACT QUANTITY MULTIPLIER MATRIX (NO MORE COPY-PASTE DOUBLE COUNTING)
                if b_qty > 0 and s_qty == 0:
                    # Single active long trade row entry calculation
                    row_pure_mtm = (live_rate_current - b_prc) * abs(b_qty)
                    adjusted_pnl_val = row_pure_mtm - r_brok
                elif s_qty > 0 and b_qty == 0:
                    # Single active short trade row entry calculation
                    row_pure_mtm = (s_prc - live_rate_current) * abs(s_qty)
                    adjusted_pnl_val = row_pure_mtm - r_brok
                else:
                    # Fallback structural proportional parsing if hybrid values exist inside a single row
                    row_pure_mtm = (live_rate_current - b_prc) * abs(b_qty - s_qty)
                    adjusted_pnl_val = row_pure_mtm - r_brok
                
                live_prices_list.append(f"{live_rate_current:,.2f}")
                floating_mtm_list.append(f"{row_pure_mtm:+,.2f} [OPEN]")
                adjusted_final_pnl_list.append(adjusted_pnl_val)
                row_visual_status_tag.append("OPEN")
                
        # Insert synchronized dynamic lists into rendering target data frames
        filtered_t6_df['Live Price'] = live_prices_list
        filtered_t6_df['Position MTM Tracking'] = floating_mtm_list
        filtered_t6_df['Net Adjusted P&L'] = adjusted_final_pnl_list
        filtered_t6_df['Internal Status Tag'] = row_visual_status_tag
        
        # Upper case normalization layers for clean UI display parameters
        filtered_t6_df['script_name'] = filtered_t6_df['script_name'].astype(str).str.upper()
        filtered_t6_df['selected_expiry'] = filtered_t6_df['selected_expiry'].astype(str).str.upper()
        
        # Formatting column data frames display headers matrix layout
        display_t6_view = filtered_t6_df[['id', 'client_name', 'week_block', 'exchange', 'script_name', 'selected_expiry', 'action_type', 'buy_qty', 'buy_price', 'sell_qty', 'sell_price', 'turnover', 'brokerage', 'Live Price', 'Position MTM Tracking', 'Net Adjusted P&L', 'Internal Status Tag']].copy()
        display_t6_view.columns = ['ID', 'Client Name', 'Week Block', 'Exchange', 'Script Name', 'Expiry', 'Action Type', 'Buy Qty', 'Buy Price', 'Sell Qty', 'Sell Price', 'Turnover', 'Brokerage', 'Live Price (₹)', 'Position MTM Tracking', 'Net Realized (₹)', 'Internal Status Tag']
        
        # 🔥 ROW CELL COLOR SWITCH MAPS WITH CONTRAST RECOVERY LOCKS
        def apply_tab6_row_formatting(row):
            if str(row['Internal Status Tag']) == 'OPEN':
                return ['background-color: #fff8e1; color: #b27a00; font-weight: bold;'] * len(row) # Amber for open row matching inventory
            else:
                return ['background-color: #f1fdf5; color: #14452b;'] * len(row) # Soft green for settled closed positions
                
        st.dataframe(
            display_t6_view.style.apply(apply_tab6_row_formatting, axis=1).format({
                "Buy Price": "{:,.2f}", "Sell Price": "{:,.2f}",
                "Turnover": "{:,.2f}", "Brokerage": "{:,.2f}", "Net Realized (₹)": "{:,.2f}"
            }),
            use_container_width=True,
            hide_index=True,
            column_order=['ID', 'Client Name', 'Week Block', 'Exchange', 'Script Name', 'Expiry', 'Action Type', 'Buy Qty', 'Buy Price', 'Sell Qty', 'Sell Price', 'Turnover', 'Brokerage', 'Live Price (₹)', 'Position MTM Tracking', 'Net Realized (₹)']
        )
        
        # 🔥 REAL TIME AUDIT SUMMARY COMPILING ENGINE
        t6_net_final_pnl = float(sum(adjusted_final_pnl_list))
        t6_brokerage_sum = float(filtered_t6_df['brokerage'].sum())
        
        t6_card_style = "pastel-green" if t6_net_final_pnl >= 0 else "pastel-red"
        t6_label = f"Total Consolidated [{selected_t6_scrip}] Realized Net Profit" if t6_net_final_pnl >= 0 else f"Total Consolidated [{selected_t6_scrip}] Realized Net Loss"
        if "SHOW ALL" in selected_t6_scrip:
            t6_label = "Total Combined Portfolio Strategy Net Evaluation"
            
        st.markdown(f"""
            <div class="metric-card {t6_card_style}" style="margin-top: 10px;">
                🎯 <b>{t6_label} (Column X Dynamic Matrix Total):</b> 
                <span style="font-size: 18px; font-weight: bold; margin-left: 10px;">₹ {t6_net_final_pnl:,.2f}</span>
                <br><small style="font-size: 11px; opacity: 0.85;">Audit Path: Live Adjusted MTM Value / Closed Trading Segments minus System Brokerage Friction (₹ {t6_brokerage_sum:,.2f})</small>
            </div>
        """, unsafe_allow_html=True)
    else:
        st.info(f"Active client [ {active_client} ] ke is week block mein filhal koi executed trades available nahi hain.")
st.write("---")
# ==========================================
# [PART_36_END]
# ==========================================
# ==========================================
# ==========================================
# [PART_37_START] - Master Destructive Wipe Out Panel
# ==========================================
st.sidebar.write("---")
st.sidebar.subheader("🚨 System Maintenance Hub")
chk_wipe_master = st.sidebar.checkbox("Confirm Master Database Purge (Warning: Destructive)", key="m_wipe_chk")
if st.sidebar.button("⚠️ WIPE COCKPIT DATA & PURGE SYSTEM LOGS", type="primary", use_container_width=True, key="master_wipe_db_btn"):
    if chk_wipe_master:
        conn = sqlite3.connect('salasar_wealth_v19_ultimate.db')
        cursor = conn.cursor()
        cursor.execute("DELETE FROM trades")
        cursor.execute("DELETE FROM cash_transactions")
        conn.commit(); conn.close()
        st.sidebar.warning("🔥 Core tables wiped clean! System reset successful.")
        st.rerun()
    else: st.sidebar.warning("⚠️ Action blocked! Check confirmation warning box first.")
# ==========================================
# [PART_37_END]
# ==========================================
import streamlit as st

# अपनी असली DB फ़ाइल का नाम यहाँ लिखें (जैसे: data.db या share_bazaar.db)
db_name = "salasar_wealth_v19_ultimate.db" 

try:
    with open(db_name, "rb") as file:
        st.download_button(
            label="Get Live DB File",
            data=file,
            file_name=db_name,
            mime="application/octet-stream"
        )
except FileNotFoundError:
    st.error("DB फ़ाइल नहीं मिली! कृपया सही नाम लिखें।")


